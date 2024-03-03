import os
import glob
import re
from collections import defaultdict, namedtuple
from operator import itemgetter
from copy import deepcopy
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

import thinkpi.operations.loader as ld
from thinkpi.operations import speed as spd
from thinkpi import logger


class WaveMeasure:
    '''Main class to handle time domain waveform measurements.
    '''

    def __init__(self, waves_path=None, port_map=None):
        '''Initializes the genereated object.

        :param waves_path: Folder path to the time domain waveforms to be processed and measured
        Files in this folder can be any type ThinkPI knows how to upload, defaults to None
        :type waves_path: str, optional
        :param port_map: Path to the port map topological .csv file,
        representing ports location in the layout database, defaults to None
        :type port_map: str, optional
        '''

        self.waves_path = waves_path
        self.port_map = port_map
        self.tran_waves = ld.Waveforms()
        self.tran_waves.path = self.waves_path
        self.heatmap_data = {}
        self.data = None
        self.time_wins = defaultdict(list)
        self.db = None
        if self.waves_path is not None:
            self.nodes, self.wave_groups = self._waveform_premeasure()
        self.filtered_wave_groups = {}

    def _waveform_premeasure(self):
        '''Generates and returns names of the probed nodes and groups of waves based on file name.

        :return: Node names and waveform groups
        :rtype: list[str], defaultdict[list]
        '''

        if self.port_map is not None:
            port_map = Path(self.port_map).read_text()

        # Organize waveforms in a dict by file name keys
        waves_by_file = defaultdict(list)
        nodes = []
        for fname in glob.glob(os.path.join(self.waves_path, '*.*')):
            if Path(fname).suffix != '.log':
                tran_waves = ld.Waveforms()
                tran_waves.load_waves(fname)
                for wave_name, wave in tran_waves.waves.items():
                    if self.port_map is not None and wave_name not in port_map:
                        continue
                    waves_by_file[os.path.basename(fname)].append(wave)
                    nodes.append(wave_name)

        # Create groups out of the waveforms
        for wave_name, waves in waves_by_file.copy().items():
            waves_by_file[wave_name] = self.tran_waves.group(*waves)
            waves_by_file[wave_name].path = self.waves_path
        
        return list(set(nodes)), waves_by_file

    def _find_nodes(self, *select_nodes):
        '''Find probe node names using wildcards if needed.

        :return: Boolean list with values representing if a node is selected or not,
        as well as node the selected node names
        :rtype: list[bool], list[str]
        '''

        found_nodes_bool = []
        found_nodes = []
        for group_node in select_nodes:
            group_node = group_node.replace('*', '.*').replace('?', '.?').replace('(', r'\(')
            match = re.compile(f"^({group_node})$")
    
            for node in self.nodes:
                if re.search(match, node):
                    found_nodes_bool.append(True)
                    found_nodes.append(node)
                else:
                    found_nodes_bool.append(False)

        return found_nodes_bool, found_nodes

    def _auto_column_width(self, writer, sheet_name):
        '''Modifies DataFrame column widths to enable full visibility of all data.

        :param df: Pandas DataDrame
        :type df: pandas.DataFrame
        :param writer: Excel writer object
        :type writer: ExcelWriter
        :param sheet_name: Name of the Excel sheet
        :type sheet_name: str
        '''

        ws = writer.sheets[sheet_name]
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=40)

        ws.column_dimensions = dim_holder
        
        '''
        for column in df:
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column) + 1
            writer.sheets[sheet_name].set_column(col_idx, col_idx, column_width)

        # Modify index column width
        index_name = '' if df.index.name is None else df.index.name
        index_width = max(df.index.astype(str).map(len).max(), len(index_name))
        writer.sheets[sheet_name].set_column(0, 0, index_width)
        '''

    def create_heatmaps(self, suffix=None):
        '''Creates colored heatmaps as an Excel file, based on the port map file.
        A single heatmap is created per a group of waveforms.

        :param suffix: Expression to add as a suffix to the report .csv name, defaults to None
        :type suffix: str, optional
        '''

        if self.port_map is None:
            return

        suffix = '' if suffix is None else f'_{suffix}'
        for group_name in self.wave_groups.keys():
            heatmap_df = pd.DataFrame(self.heatmap_data[group_name])
            heatmap_df.set_index('Node Name', drop=True, inplace=True)
            self.tran_waves.heatmap_vec(self.port_map, heatmap_df,
                                        heatmap_fname=f'heatmap_{group_name}{suffix}.xlsx'
                                    )

    def create_node_report(self, select_nodes=None, suffix=None):
        '''Creates measurement report for each node and groups.

        :param select_nodes: Expression to select the nodes to be reported.
        Wildcards can be used if needed. If not provided all nodes are reported, defaults to None
        :type select_nodes: str, optional
        :param suffix: Expression to add as a suffix to the report .csv name, defaults to None
        :type suffix: str, optional
        '''

        suffix = '' if suffix is None else f'_{suffix}'

        if select_nodes is None:
            nodes_to_use, nodes = [True]*len(self.nodes), self.nodes
        else:
            if isinstance(select_nodes, str):
                select_nodes = [select_nodes]
            nodes_to_use, nodes = self._find_nodes(*select_nodes)

        # Add max, min, and avg to the collected data
        data = deepcopy(self.data)
        for measure, groups in self.data.items():
            for group in groups.keys():
                data[measure][group] = list(np.array(self.data[measure][group])[nodes_to_use])
                data[measure][group] += ['', np.max(data[measure][group]),
                                            np.min(data[measure][group]),
                                            np.mean(data[measure][group])
                                    ]
        
        # Save data in multiple measurement type tabs in an Excel file
        with pd.ExcelWriter(os.path.join(self.waves_path, 'reports', f'summary_nodes{suffix}.xlsx')) as writer:
            for measure, group in data.items():
                data_df = pd.DataFrame(group, index=nodes + ['', 'Max', 'Min', 'Avg'])
                data_df.to_excel(writer, sheet_name=measure)
                self._auto_column_width(writer, measure)

            # Add last sheet with time windows if available
            if self.time_wins:
                time_wins = {}
                for group_name, time_win in self.time_wins.items():
                    time_wins[group_name] = list(np.array(time_win)[nodes_to_use])
                time_win_df = pd.DataFrame(time_wins, index=nodes)
                time_win_df.to_excel(writer, sheet_name='settle_time_windows')
                self._auto_column_width(writer, 'settle_time_windows')

    def create_group_report(self, *group_nodes, suffix=None):
        '''Creates measurement report for the requested groups.

        :param suffix: Expression to add as a suffix to the report .csv name, defaults to None
        :type suffix: str, optional
        :param group_nodes: Expressions to select the groups to be reported.
        Wildcards can be used if needed
        :type group_nodes: str, optional
        '''

        suffix = '' if suffix is None else f'_{suffix}'

        if not group_nodes:
            logger.warning('Must provide at least one group.\nGroup report generation has failed.')
            return

        table = defaultdict(dict)
        worst_min_nodes = defaultdict(list)
        worst_max_nodes = defaultdict(list)
        for group_node in group_nodes:
            group_title = group_node.replace('*', '').replace('?', '')
            report_nodes, node_names = self._find_nodes(group_node)

            if not node_names:
                logger.warning(f'Group {group_node} cannot be found')
                continue
            
            for measure, group_data in self.data.items():
                table[group_title][measure] = []
                groups = []
                for group, data in group_data.items():
                    group_nodes = [(val, node_name, group)
                                    for val, node_name in
                                        zip(np.array(data)[report_nodes], node_names)
                                ]
                    if 'min' in measure:
                        table[group_title][measure].append(min(group_nodes, key=itemgetter(0))[:2])
                        worst_min_nodes[measure] += group_nodes
                    else:
                        table[group_title][measure].append(max(group_nodes, key=itemgetter(0))[:2])
                        worst_max_nodes[measure] += group_nodes
                    
                    groups.append(group)

        # Save worst waveforms as .csv files
        for measure, group_nodes in worst_min_nodes.items():
            _, node_name, group_name = min(group_nodes, key=itemgetter(0))
            self.wave_groups[group_name].waves[node_name].save(
                                                    format_type='csv',
                                                    fname=f'{node_name}_{measure}{suffix}',
                                                    subckt_name=f'{node_name}_{measure}'
                                                )
        for measure, group_nodes in worst_max_nodes.items():
            _, node_name, group_name = max(group_nodes, key=itemgetter(0))
            wave_to_save = self.wave_groups[group_name].waves[node_name]
            if 'settle' in measure:
                wave_to_save = wave_to_save.clip(wave_to_save.results['pk2pk'][0][0],
                                                wave_to_save.results['pk2pk'][0][1]).shift()
            wave_to_save.save(
                                format_type='csv',
                                fname=f'{node_name}_{measure}{suffix}',
                                subckt_name=f'{node_name}_{measure}'
                            )

        # Save data in multiple node type tabs in an Excel file
        with pd.ExcelWriter(os.path.join(self.waves_path, 'reports', f'summary_groups{suffix}.xlsx')) as writer:
            for group_title, measure in table.items():
                data_df = pd.DataFrame(measure, index=groups)
                data_df.to_excel(writer, sheet_name=group_title)
                self._auto_column_width(writer, group_title)

    def plot_filtered(self, suffix=None, num_waves=5):
        '''Producing a series of overlayed unfiltered and filtered plots.

        :param suffix: Expression to add as a suffix to the .html name, defaults to None
        :type suffix: str, optional
        :param num_waves: Number of waveforms to plot, defaults to 5
        :type num_waves: int, optional
        '''

        suffix = '' if suffix is None else f'_{suffix}'

        for group_name, wave_group in self.wave_groups.items():
            plots = []
            for wave, filtered_wave in zip(list(wave_group.waves.values())[:num_waves],
                                           list(self.filtered_wave_groups[group_name].waves.values())[:num_waves]):
                plots.append([wave, filtered_wave])
            self.tran_waves.plot_overlay(plots, out_file=f"{group_name.split('.')[0]}{suffix}")
        
    def filter_waveform_measure(self, filter, tstart=None, tend=None, unit='V'):
        '''Creates measurement data by applying a given filter and performing the defined measurements.
        If time window is not provided, automatic detection of the settling region takes place.

        :param filter: Filter object
        :type filter: :class:`filters.Filter()`
        :param tstart: start time window to apply the filter, defaults to None
        :type tstart: float, optional
        :param tend: start time window to apply the filter, defaults to None
        :type tend: float, optional
        :param unit: The unit of the waveforms, defaults to 'V'
        :type unit: str, optional
        '''

        # Reset data structures
        self.filtered_wave_groups = {}
        self.heatmap_data = {}
        self.time_wins = defaultdict(list)

        auto_detect = True if tstart is None and tend is None else False

        freq = filter._freq_descr().replace('.', 'p')
        measures = [f'min_{unit}_{freq}',
                    f'max_{unit}_{freq}',
                    f'pk2pk_m{unit}_{freq}'
                ]

        self.data = {measure: defaultdict(list) for measure in measures}
        for group_name, group in self.wave_groups.items():
            if auto_detect:
                tstart, tend = [], []
                time_windows = group.detect_settle()
                for (ts, te) in time_windows.values():
                    tstart.append(ts)
                    tend.append(te)
            
            group = group.filt(filter, tstart, tend)
            self.filtered_wave_groups[group_name] = deepcopy(group)
            # Revert waveform names to original values
            for wave_name, wave in group.waves.copy().items():
                prev_name = '_'.join(wave_name.split('_')[:-1])
                wave.wave_name = prev_name
                group.waves[prev_name] = wave
                del group.waves[wave_name]
            results = {measure:result for measure, result in zip(measures,
                                                            [group.minimum(),
                                                            group.maximum(),
                                                            group.pk2pk(),
                                                        ]
                                                    )
                    }

            self.heatmap_data[group_name] = {measure: [] for measure in measures + ['Node Name']}
            for node in self.nodes:
                self.heatmap_data[group_name]['Node Name'].append(node)
                if auto_detect:
                    self.time_wins[group_name].append(f'tsettle: {time_windows[node][0]*1e9:.3f} '
                                                    f'to {time_windows[node][1]*1e9:.3f} ns'
                                                )
                else:
                    self.time_wins[group_name].append(f'tsettle: {tstart*1e9:.3f} '
                                                    f'to {tend*1e9:.3f} ns'
                                                )

                for measure in measures:
                    if 'pk2pk_m' in measure:
                        result = results[measure][node][1]*1e3
                    else:
                        result = results[measure][node]
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)    

    def voltage_waveform_measure_HF(self, tstart=None, tend=None):
        '''Measures high frequency PDN response to certain stimuli.
        This includes global minimum, maximum, and peak to peak voltages,
        as well as, minimum, maximum, and peak to peak voltages in the settling
        regions of the PDN response. If parameters are not provided,
        automatic detection takes place.

        :param tstart: Starting time window of the settling region.
        If both 'start' and 'tend are not provided, automatic detection is performed,
        defaults to None
        :type tstart: float, optional
        :param tend: Ending time window of the settling region.
        If both 'tstart' and 'tend are not provided, automatic detection is performed,
        defaults to None
        :type tend: float, optional
        '''

        auto_detect = True if tstart is None and tend is None else False

        measures = ['Vmin_global_V', 'Vmax_global_V', 'Vp2p_global_mV',
                    'Vavg_global_V', 'Vmin_settle_V', 'Vmax_settle_V',
                    'Vp2p_settle_mV', 'Vavg_settle_V'
                ]

        self.data = {measure: defaultdict(list) for measure in measures}
        logger.info('Processing the following cases:')
        for group_name, group in self.wave_groups.items():
            logger.info(f'\t{group_name}')
            self.heatmap_data[group_name] = {measure: [] for measure in measures + ['Node Name']}
            if auto_detect:
                tstart, tend = [], []
                time_windows = group.detect_settle()
                time_windows = {node: ((ts + te)/2, te) for node, (ts, te) in time_windows.items()}
                for (ts, te) in time_windows.values():
                    tstart.append(ts)
                    tend.append(te)
                    
            results = {measure:result for measure, result in zip(measures, [group.minimum(),
                                                                group.maximum(),
                                                                group.pk2pk(),
                                                                group.average(),
                                                                group.minimum(tstart, tend),
                                                                group.maximum(tstart, tend),
                                                                group.pk2pk(tstart, tend),
                                                                group.average(tstart, tend)])
                    }           

            for node in self.nodes:
                self.heatmap_data[group_name]['Node Name'].append(node)
                if auto_detect:
                    self.time_wins[group_name].append(f'tsettle: {time_windows[node][0]*1e9:.3f} '
                                                    f'to {time_windows[node][1]*1e9:.3f} ns'
                                                )
                else:
                    self.time_wins[group_name].append(f'tsettle: {tstart*1e9:.3f} '
                                                    f'to {tend*1e9:.3f} ns'
                                                )
                for measure in measures:
                    if measure == 'Vp2p_global_mV' or measure == 'Vp2p_settle_mV':
                        result = results[measure][node][1]*1e3
                    else:
                        result = results[measure][node]
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)

    def _auto_voltage_waveform_measure_LF(self):
        '''Automatic detection and measurement
        of a low frequency voltage PDN response. This includes first, second,
        and third minimum and droops measurement, as well as, first, second,
        and third maximum and overshoots measurements.
        '''

        measures = ['min_1st_V', 'droop_1st_mV', 'min_2nd_V', 'droop_2nd_mV',
                    'min_3rd_V', 'droop_3rd_mV', 'max_1st_V', 'over_1st_mV',
                    'max_2nd_V', 'over_2nd_mV', 'max_3rd_V', 'over_3rd_mV',
                    'DC_unloaded_V', 'DC_loaded_V', 'Vmin_global_V',
                    'Vmax_global_V', 'Vp2p_global_mV']

        self.data = {measure: defaultdict(list) for measure in measures}
        for group_name, group in self.wave_groups.items():
            group_droops = group.find_droops()
            group_overs = group.find_overshoots()
            group_min = group.minimum()
            group_max = group.maximum()
            group_pk2pk = group.pk2pk()

            self.heatmap_data[group_name] = {measure: [] for measure in measures + ['Node Name']}
            for node in self.nodes:
                self.heatmap_data[group_name]['Node Name'].append(node)

                for measure, result in group_droops[node].items():
                    self.heatmap_data[group_name][measure].append(result[1])
                    self.data[measure][group_name].append(result[1])
    
                for measure, result in group_overs[node].items():
                    self.heatmap_data[group_name][measure].append(result[1])
                    self.data[measure][group_name].append(result[1])

                for measure, result in zip(measures[-3:],
                                            [group_min[node],
                                             group_max[node],
                                             group_pk2pk[node][1]*1e3]):
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)

        # Delete unused measures
        for group_name in self.heatmap_data.copy().keys():
            for measure in measures:
                if not self.heatmap_data[group_name][measure]:
                    del self.heatmap_data[group_name][measure]
        for measure in measures:
            if not self.data[measure]:
                del self.data[measure]

    def voltage_waveform_measure_LF(self, tDC_unload=None,
                                    tfirst_droop=None,
                                    tsecond_droop=None,
                                    tthird_droop=None,
                                    tDC_load=None,
                                    tfirst_over=None,
                                    tsecond_over=None,
                                    tthird_over=None):
        '''Manual measurement of low frequency voltage PDN response,
        based on input parameters. This includes first, second,
        and third minimum and droops measurement, as well as,
        first, second, and third maximum and overshoots measurements.
        If None of the parameters are provided, automatic detection takes place.

        :param tDC_unload: Point in time at which the system is
        constantly unloaded (or constantly loaded with a small load), defaults to None
        :type tDC_unload: float, optional
        :param tfirst_droop: Point in time in the vicinity of the first droop, defaults to None
        :type tfirst_droop: float, optional
        :param tsecond_droop: Point in time in the vicinity of the second droop, defaults to None
        :type tsecond_droop: float, optional
        :param tthird_droop: Point in time in the vicinity of the third droop, defaults to None
        :type tthird_droop: float, optional
        :param tDC_load: Point in time at which the system is constantly loaded,
        defaults to None
        :type tDC_load: float, optional
        :param tfirst_over: Point in time in the vicinity of the first overshoot, defaults to None
        :type tfirst_over: float, optional
        :param tsecond_over: Point in time in the vicinity of the second overshoot, defaults to None
        :type tsecond_over: float, optional
        :param tthird_over: Point in time in the vicinity of the third overshoot, defaults to None
        :type tthird_over: float, optional
        '''

        auto_detect = True
        measures = []
        measure_times = {'min_1st_V': tfirst_droop, 'droop_1st_mV': tfirst_droop,
                        'min_2nd_V': tsecond_droop, 'droop_2nd_mV': tsecond_droop,
                        'min_3rd_V': tthird_droop, 'droop_3rd_mV': tthird_droop,
                        'max_1st_V': tfirst_over, 'over_1st_mV': tfirst_over,
                        'max_2nd_V': tsecond_over, 'over_2nd_mV': tsecond_over,
                        'max_3rd_V': tthird_over, 'over_3rd_mV': tthird_over}

        for tmeasure_type, tmsr in measure_times.items():
            if tmsr is not None:
                auto_detect = False
                measures.append(tmeasure_type)
        if tDC_unload is not None:
            auto_detect = False
            measures.append('DC_unloaded_V')
        if tDC_load is not None:
            auto_detect = False
            measures.append('DC_loaded_V')

        if auto_detect:
            self._auto_voltage_waveform_measure_LF()
            return
                    
        self.data = {measure: defaultdict(list) for measure in (measures
                                                                + ['Vmin_global_V',
                                                                    'Vmax_global_V',
                                                                    'Vp2p_global_mV']
                                                            )
                }
        for group_name, group in self.wave_groups.items():
            taxis = group.waves[list(group.waves.keys())[0]].data.x
            group_ops = []
            for measure in measures:
                if '_unloaded' in measure:
                    group_ops.append(group.y_at_x(tDC_unload))
                    continue
                if '_loaded' in measure:
                    group_ops.append(group.y_at_x(tDC_load))
                    continue
                # Find index
                idx = np.where(taxis >= measure_times[measure])[0][0]
                tstart = taxis[int(np.ceil(idx*0.95))]
                tend = taxis[int(np.ceil(idx*1.05))]
                if 'min' in measure or 'droop' in measure:
                    group_ops.append(group.minimum(tstart, tend))
                elif 'max' in measure or 'over' in measure:
                    group_ops.append(group.maximum(tstart, tend))
            results = {measure: result for measure, result in zip(measures, group_ops)}
            # Measure global minimum, maximum, and peak-to-peak
            group_min = group.minimum()
            group_max = group.maximum()
            group_pk2pk = group.pk2pk()

            self.heatmap_data[group_name] = {measure: []
                                             for measure in (measures
                                                            + ['Node Name',
                                                               'Vmin_global_V',
                                                                'Vmax_global_V',
                                                                'Vp2p_global_mV']
                                                        )
                                        }
            for node in self.nodes:
                self.heatmap_data[group_name]['Node Name'].append(node)
                for measure in measures:
                    result = results[measure][node]
                    if not auto_detect:
                        if 'DC_unloaded_V' in measures and 'droop' in measure:
                            result = (results['DC_unloaded_V'][node] - result)*1e3
                        if 'DC_loaded_V' in measures and 'over' in measure:
                            result = (result - results['DC_loaded_V'][node])*1e3
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)

                for measure, result in zip(['Vmin_global_V',
                                            'Vmax_global_V',
                                            'Vp2p_global_mV'],
                                            [group_min[node],
                                            group_max[node],
                                            group_pk2pk[node][1]*1e3]):
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)

    def current_waveform_measure(self, tmin=(None, None),
                                    tmax=(None, None),
                                    tslew=(None, None),
                                    tpeak=(None, None),
                                    slew_perc=90
                                ):
        '''Measures current waveforms, based on the provided parameters.
        This includes minimum, maximum, peak, average, and slew rate current measurements.
        If Parameters are not provided automatic detection takes place.
        Note that automatic detection can be mixed with manually providing the time windows.

        :param tmin: Time window for the minimum region, defaults to (None, None)
        :type tmin: tuple, optional
        :param tmax: Time window for the maximum region, defaults to (None, None)
        :type tmax: tuple, optional
        :param tslew: Time window for the slew region, defaults to (None, None)
        :type tslew: tuple, optional
        :param tpeak: Time window for the peak region, defaults to (None, None)
        :type tpeak: tuple, optional
        :param slew_perc: Percent to measure the waveform slew rate.
        For example, 90 means to measure slew rate between 90% and 10% of the waveform,
        defaults to 90
        :type slew_perc: int, optional
        '''

        auto_detect_tmin = True if tmin == (None, None) else False
        auto_detect_tmax = True if tmax == (None, None) else False
        auto_detect_tslew = True if tslew == (None, None) else False

        measures = ['Imin_A', 'Imax_A', 'Ipeak_A', 'Imin_avg_A',
                    'Imax_avg_A', 'Islew_A_usec'
                ]

        self.data = {measure: defaultdict(list) for measure in measures}
        for group_name, group in self.wave_groups.items():
            if auto_detect_tmin or auto_detect_tmax or auto_detect_tslew:
                if auto_detect_tmin:
                    tmin = ([], [])
                if auto_detect_tmax:
                    tmax = ([], [])
                if auto_detect_tslew:
                    tslew = ([], [])
                time_windows = group.detect_change_points(plot_fig=False,
                                                            slew_perc=slew_perc)
                for twins in time_windows.values():
                    if auto_detect_tmin:
                        tmin[0].append(twins[0][0])
                        tmin[1].append(twins[0][1])
                    if auto_detect_tmax:
                        tmax[0].append(twins[1][0])
                        tmax[1].append(twins[1][1])
                    if auto_detect_tslew:
                        tslew[0].append(twins[-1][1])
                        tslew[1].append(twins[-1][2])
            
            results = {measure: result for measure, result in zip(measures,
                                                                    [group.minimum(tmin[0], tmin[1]),
                                                                    group.maximum(tmax[0], tmax[1]),
                                                                    group.maximum(tpeak[0], tpeak[1]),
                                                                    group.average(tmin[0], tmin[1]),
                                                                    group.average(tmax[0], tmax[1]),
                                                                    group.slew_rate(tslew[0], tslew[1])
                                                                ]
                                                            )
                    }

            self.heatmap_data[group_name] = {measure: [] for measure in measures + ['Node Name']}
            for node in self.nodes:
                self.heatmap_data[group_name]['Node Name'].append(node)

                detected_windows = ''
                if auto_detect_tmin:
                    detected_windows += (f'tmin: {time_windows[node][0][0]*1e9:.3f} '
                                        f'to {time_windows[node][0][1]*1e9:.3f} ns | '
                                    )
                else:
                    detected_windows += (f'tmin: {tmin[0]*1e9:.3f} '
                                        f'to {tmin[0]*1e9:.3f} ns | '
                                    )
                if auto_detect_tmax:
                    detected_windows += (f'tmax: {time_windows[node][1][0]*1e9:.3f} '
                                        f'to {time_windows[node][1][1]*1e9:.3f} ns | '
                                    )
                else:
                    detected_windows += (f'tmax: {tmax[0]*1e9:.3f} '
                                        f'to {tmax[0]*1e9:.3f} ns | '
                                    )
                if auto_detect_tslew:
                    detected_windows += (f'tslew: {time_windows[node][2][1]*1e9:.3f} '
                                        f'to {time_windows[node][-1][-1]*1e9:.3f} ns'
                                    )
                else:
                    detected_windows += (f'tslew: {tslew[0]*1e9:.3f} '
                                        f'to {tslew[0]*1e9:.3f} ns'
                                    )
                self.time_wins[group_name].append(detected_windows)
                
                for measure in measures:
                    result = results[measure][node]
                    if result is not None and measure == 'Islew_A_usec':
                        result = result/1e6
                    self.heatmap_data[group_name][measure].append(result)
                    self.data[measure][group_name].append(result)

    def export_port_map(self, fname, db_fname):
        '''Create .csv port map file, representing the topological
        location of the ports in the layout database.

        :param fname: File name of the created .csv file
        :type fname: str
        :param db_fname: Layout database (.spd) file name
        :type db_fname: str
        '''

        grid = GridCreator(db_fname)
        grid.export_port_map(fname, self)

    def post_process(self, fname, tstart_clip=None, tend_clip=None,
                    shift_dt=None, tresample=1e-12, decimals=12):
        '''Applies post processing operations on a given waveform or a folder of waveforms.
        This includes, clipping, shifting, and resampling.
        The new generated waveform(s) are saved as .csv files.

        :param fname: Path to a folder containing waveforms or a file name of a single waveform
        :type fname: str
        :param tstart_clip: Starting time clipping window, defaults to None
        :type tstart_clip: float, optional
        :param tend_clip: Ending time clipping window, defaults to None
        :type tend_clip: float, optional
        :param shift_dt: Delta shift time in Seconds.
        If not provided, waveform is shifted to time zero, defaults to None
        :type shift_dt: float, optional
        :param tresample: Sampling step in Seconds, defaults to 1e-12
        :type tresample: float, optional
        '''

        waves = ld.Waveforms()
        waves.load_waves(fname)

        proc_waves = waves.clip(tstart_clip, tend_clip).shift(shift_dt).resample(tresample, verbose=False)

        for wave in proc_waves.waves.values():
            wave.save('csv', suffix=f"{wave.file_name.split('.')[0]}", decimals=decimals)

    def plot(self):
        '''Generating stacked plots of all imported nodes per file.
        '''

        for from_file, waves in self.wave_groups.items():
            waves.plot_stack(out_file=Path(from_file).stem)


class GridCreator:
    '''Class to handle port map creation based on layout .spd files.
    '''

    def __init__(self, db):
        '''Initializes the generated objects.

        :param db: Name of a layout database or a database object
        :type db: str or :class:`speed.Database()`
        '''

        if isinstance(db, str):
            db = spd.Database(db)
            db.load_flags = {'layers': False,
                            'nets': False,
                            'nodes': True,
                            'ports': True,
                            'shapes': False,
                            'padstacks': False,
                            'vias': False,
                            'components': False,
                            'traces': False,
                            'sinks': False,
                            'vrms': False,
                            'plots': False
                        }
            db.load_data()
        self.db = db
        self.port_name_conv = None

    def export_port_map(self, fname, measure):
        '''Exports probed node to port map file for selected probed nodes.
        The mapping is performed by checking if the port name is part of a the node name.

        :param fname: Name of the .csv mapping file to be exported
        :type fname: str
        :param measure: Probed nodes for which the mapping is created
        :type measure: list
        '''

        # Find matches between waveform names and port names
        name_conv = {}
        logger.info('\nMatched ports to simulation nodes:')
        for port_name in self.db.port_names(verbose=False):
            name_conv[port_name] = ''
            for node in measure.nodes:
                if port_name in node:
                    name_conv[port_name] = node
                    logger.info(f'\t{node} --> {port_name}')
                    break

        # Find unmatched nodes
        matched_nodes = list(name_conv.values())
        unmatched_nodes = []
        for node in measure.nodes:
            if node not in matched_nodes:
                unmatched_nodes.append(node)

        map_df = pd.DataFrame([(key, val) for key, val in name_conv.items()])
        map_df = pd.concat([map_df, pd.DataFrame(unmatched_nodes)], ignore_index=True, axis=1)
        map_df.to_csv(fname, header=False, index=False)
        self.port_conv_fname = fname
    
    def coord_conv(self, xls_x2, xls_y2):
        '''Calculates conversion parameters relating layout database and .csv file coordinate systems.
        These parameters are later used to convert between layout database port locations
        and .csv file locations.

        :param xls_x2: Coordinate of the X bottom right row number in .csv file
        :type xls_x2: int
        :param xls_y2: Coordinate of the Y bottom right column number in .csv file
        :type xls_y2: int
        :return: Coordinate system conversion parameters
        :rtype: namedtuple
        '''

        Trans = namedtuple('Trans', 'tfx sfx tfy sfy')
        return Trans(tfx=(xls_x2*self.db.db_x_bot_left - self.db.db_x_top_right)/(self.db.db_x_bot_left - self.db.db_x_top_right),
                     sfx=(xls_x2 - 1)/(self.db.db_x_top_right - self.db.db_x_bot_left),
                     tfy=-(xls_y2*self.db.db_y_bot_left - self.db.db_y_top_right)/(self.db.db_y_bot_left - self.db.db_y_top_right),
                     sfy=-(xls_y2 - 1)/(self.db.db_y_top_right - self.db.db_y_bot_left)
                    )

    def port_map(self, xls_x2, xls_y2, elems):
        '''Creates a matrix representing .csv file.
        The values in the matrix are the probed node names and their location
        is caluclated based on the coordinate system transformation. 

        :param xls_x2: Coordinate of the X bottom right row number in .csv file
        :type xls_x2: int
        :param xls_y2: Coordinate of the Y bottom right column number in .csv file
        :type xls_y2: int
        :param elems: Elements to use for the construction of the grid.
        Currently, will work only with `self.ports` and `self.sinks`. 
        :type elems: dict[:class:`speed.Port()`|:class:`speed.Sink()`]
        :return: Matrix representing .csv file
        :rtype: numpy.array
        '''

        grid = np.empty(shape=(xls_y2+1, xls_x2+1), dtype=object)
        trans = self.coord_conv(xls_x2, xls_y2)

        for port_name, node_name in self.port_name_conv.items():
            x_grid = round(trans.tfx + trans.sfx*elems[port_name].x_center)
            y_grid = round(trans.tfy + trans.sfy*elems[port_name].y_center)
            if grid[y_grid, x_grid] is not None:
                pass #return None
            else:
                grid[y_grid, x_grid] = node_name

        return grid

    def compress_grid(self, min_val, max_val, axis):
        '''Removes blank raws and columns in order to reduce the size of the matrix.
        This is done such that the resulting .csv file will be as small as possible.

        :param min_val: Minimum value of a raw or column depedning on the axis
        :type min_val: int
        :param max_val: Maximum value of a raw or column depedning on the axis
        :type max_val: int
        :param axis: Axis along which the reduction is performed. Can only accepts 'x' or 'y'
        :type axis: str
        :return: Thew new maximum value
        :rtype: int
        '''

        mid_val = int((min_val + max_val)/2)
        while True:
            if axis == 'x':
                grid = self.port_map(mid_val, max_val)
            else:
                grid = self.port_map(max_val, mid_val)
            if grid is None:
                min_val = mid_val
            else:
                max_val = mid_val
            mid_val = int((min_val + max_val)/2)
            if min_val == mid_val or max_val == mid_val:
                return max_val

    def create_grid(self, fname, port_conv_fname=None, use_sinks=False):
        '''Creates a .csv file with node names placed according to
        their locations in the layout database.

        :param fname: File name of the created .csv map file
        :type fname: str
        :param port_conv_fname: File name of the ports to nodes conversion map
        :type port_conv_fname: str
        :param use_sinks: If True, use sink locations and names to create the grid.
        Otherwise ports are used, defaults to False
        :type use_sinks: bool, optional
        '''

        # Read mapping file and convert to a dict
        if use_sinks:
            elems = self.db.sinks
        else:
            elems = self.db.ports

        if port_conv_fname is None:
            self.port_name_conv = {port_name: port_name for port_name in elems.keys()}
        else:
            self.port_name_conv = pd.read_csv(port_conv_fname,
                                                index_col=0,
                                                usecols=[0, 1],
                                                header=None)
            self.port_name_conv = self.port_name_conv.dropna()
            self.port_name_conv = self.port_name_conv.to_dict()[1]
        
        xls_x2 = 1000
        xls_y2 = 1000
        while True:
            grid = self.port_map(xls_x2, xls_y2, elems)
            if grid is None:
                xls_x2 *= 2
                xls_y2 *= 2
            else:
                break
        
        grid_df = pd.DataFrame(grid)
        grid_df = grid_df.dropna(how='all', axis=0).reset_index(drop=True)
        grid_df = grid_df.dropna(how='all', axis=1)
        grid_df.columns = list(range(len(grid_df.columns)))
        grid_df.to_csv(fname) #, index=False, header=False)
