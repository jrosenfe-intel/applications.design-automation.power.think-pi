import os
from pathlib import Path
from copy import deepcopy
from collections import defaultdict, namedtuple
from operator import itemgetter
from itertools import cycle
from importlib import reload 
from difflib import get_close_matches

from scipy import signal
from scipy.interpolate import interp1d
from scipy.ndimage import median_filter
from scipy.signal import blackman, hann, hamming, bartlett
from scipy.fft import fft, ifft

import pandas as pd
import numpy as np
from bokeh.io import output_file, show
from bokeh.layouts import column
from bokeh.palettes import Category20, Category10

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule

import skrf as rf
from skrf.network import Network

import matplotlib.pyplot as plt

from thinkpi.operations.plotting import Plotter
from thinkpi.operations.speed import Database
import thinkpi.waveforms.primitives as prim

import thinkpi.operations.loader as ld
from thinkpi import DataVector, logger


class Ops:
    '''This class defines possible operations on a single individual waveform.
    '''

    def __init__(self):

        self.plt = Plotter()

    @property
    def start(self):

        return self.data.x[0]

    @property
    def end(self):

        return self.data.x[-1]

    def _clip(self, tstart=None, tend=None):
        
        if tstart is None and tend is None:
            return self.data.x, self.data.y

        inter_func = interp1d(self.data.x, self.data.y)
        new_x = self.data.x
        try:
            if tstart is None:
                idx_start = 0
            else:
                idx_start = np.where(new_x >= tstart)[0][0]
                new_x = np.insert(new_x, idx_start, tstart)
            if tend is None:
                idx_end = -1
            else:
                idx_end = np.where(new_x <= tend)[0][-1]
                new_x = np.insert(new_x, idx_end, tend)
        except IndexError: # Catches if tstart or tend is larger than the waveform
            return self.data.x, self.data.y
        new_x = new_x[idx_start:] if idx_end == -1 else new_x[idx_start:idx_end+1]
        new_y = inter_func(new_x)

        return new_x, new_y

    def clip(self, tstart=None, tend=None):
        '''Truncate a waveform at defined start and end times

        :param tstart: start time of the truncation window in seconds, defaults to None
        :type tstart: float, None, optional
        :param tend: end time of the truncation window in seconds, defaults to None
        :type tend: float, None, optional
        :return: Clipped waveform
        :rtype: :class:`primitives.Wave()`
        '''

        x_clip, y_clip = self._clip(tstart, tend)
        history = self.history + [f'Clip: {x_clip[0]} to {x_clip[-1]}']

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=x_clip, y=y_clip,
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_clipped{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=history
                              )
                    )

    def abs(self):
        '''Calculate the absolute value of a waveform, element-wise, or a magnitude if the waveform is complex

        :return: Waveform magnitude
        :rtype: :class:`primitives.Wave()`
        '''
        
        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=self.data.x, y=np.abs(self.data.y),
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_abs{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=self.history
                              )
                    )

    def angle(self):
        '''Calculate the phase of a complex waveform element-wise

        :return: Waveform phase
        :rtype: :class:`primitives.Wave()`
        '''

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=self.data.x, y=np.angle(self.data.y),
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_angle{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=self.history
                              )
                    )

    def real(self):
        '''Calculate the real portion of a complex waveform element-wise

        :return: Real part vector
        :rtype: :class:`primitives.Wave()`
        '''

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=self.data.x, y=np.real(self.data.y),
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_real{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=self.history
                              )
                    )

    def imag(self):
        '''Calculate the imaginary portion of a complex waveform element-wise

        :return: Imaginary part vector
        :rtype: :class:`primitives.Wave()`
        '''

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=self.data.x, y=np.imag(self.data.y),
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_imag{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=self.history
                              )
                    )

    def plot(self, wave_name=None, x_scale='u', y_scale='', xaxis_type='linear',
             yaxis_type='linear', clr='navy', stretch_mode=True):

        if wave_name is None:
            wave_name = self.wave_name
            for invalid_char in ['/', '\\', '(', ')', '<', '>']:
                wave_name = wave_name.replace(invalid_char, '')
            wave_name = wave_name.split(' ')[0]
        output_file(os.path.join(self.path, 'plots', f'{wave_name}.html'))
        try:
            show(self.plt.single_plot(self, x_scale=x_scale, y_scale=y_scale,
                                      xaxis_type=xaxis_type, yaxis_type=yaxis_type,
                                      clr=clr, stretch_mode=stretch_mode
                                     )
                )
        except OSError:
            logger.warning(f'Failed to save plot using the waveform name {wave_name}.'
                  f'\nProvide a legal file name in the plot command or change waveform name.')

    def x_at_y(self, y):
        '''Finds x value at a given y value.

        :param y: Vertical axis value
        :type y: float
        :return: The corresponding horizontal x-axis value
        :rtype: float
        '''

        y_line = np.array([y]*len(self.data.y))
        idx = np.argwhere(np.diff(np.sign(self.data.y - y_line))).flatten()

        self.results['x_at_y'] = (self.data.x[idx[0]], y)

        return self.results['x_at_y'][0]

    def y_at_x(self, x):
        '''Finds y value at a given x value.

        :param x: Horizontal axis value
        :type x: float
        :return: The corresponding vertical y-axis value
        :rtype: float
        '''
        
        wave_func = interp1d(self.data.x, self.data.y)
        self.results['y_at_x'] = (x, float(wave_func(x)))

        return self.results['y_at_x'][1]

    def maximum(self, tstart=None, tend=None):
        '''Find the maximum value of a waveform and the time/frequency it occrs at between starting and ending times

        :param tstart: starting time in seconds, defaults to None
        :type tstart: float, optional
        :param tend: ending time in seconds, defaults to None
        :type tend: float, optional
        :return: A tuple of time and maximum value
        :rtype: tuple[float, float]
        '''

        try:
            _, y_clip = self._clip(tstart, tend)
            self.results['max'] = np.max(y_clip)
            self.results['max'] = (self.data.x[np.where(self.data.y == np.max(y_clip))[0][0]],
                                    np.max(y_clip))
            self.results['max'][1]
        except IndexError:
            return None

        return self.results['max'][1]

    def minimum(self, tstart=None, tend=None):
        '''Find the minimum value of a waveform and the time/frequency it occrs at between starting and ending times

        :param tstart: starting time in seconds, defaults to None
        :type tstart: float, optional
        :param tend: ending time in seconds, defaults to None
        :type tend: float, optional
        :return: A tuple of time and minimum value
        :rtype: tuple[float, float]
        '''

        try:
            _, y_clip = self._clip(tstart, tend)
            self.results['min'] = (self.data.x[np.where(self.data.y == np.min(y_clip))[0][0]], np.min(y_clip))
        except IndexError:
            return None

        return self.results['min'][1]

    def pk2pk(self, tstart=None, tend=None):
        '''Find peak to peak value of a waveform between starting and ending times

        :param tstart: starting time in seconds, defaults to None
        :type tstart: float, optional
        :param tend: ending time in seconds, defaults to None
        :type tend: float, optional
        :return: Peak to peak value
        :rtype: float
        '''

        try:
            self.results['pk2pk'] = ((tstart, tend), self.maximum(tstart, tend) - self.minimum(tstart, tend))
        except IndexError:
            return None

        return self.results['pk2pk']

    def mid_point(self, tstart=None, tend=None):

        try:
            self.results['mid_point'] = (self.maximum(tstart, tend) + self.minimum(tstart, tend))/2
        except IndexError:
            return None

        return self.results['mid_point']

    def average(self, tstart=None, tend=None):

        try:
            _, y_clip = self._clip(tstart, tend)
            self.results['avg'] = np.mean(y_clip)
        except IndexError:
            return None
        
        return self.results['avg']

    def _find_DC_unloaded(self, tstart=None, tend=None, eps=100):

        x_clip, y_clip = self._clip(tstart, tend)

        # Find DC unloaded
        try:
            min_idx = np.where(x_clip >= self.results['min_1st_V'][0])[0][0]
        except KeyError:
            return

        for idx in range(min_idx, 1, -1):
            if (x_clip[idx] - x_clip[idx-1]) == 0:
                continue
            slope = abs((y_clip[idx] - y_clip[idx-1])/(x_clip[idx] - x_clip[idx-1]))
            if slope < eps:
                self.results['DC_unloaded_V'] = (x_clip[idx], y_clip[idx])
                break

    def _find_DC_loaded(self, tstart=None, tend=None, eps=100):

        x_clip, y_clip = self._clip(tstart, tend)

        # Find DC loaded
        try:
            max_idx = np.where(x_clip >= self.results['max_1st_V'][0])[0][0]
        except KeyError:
            return

        for idx in range(max_idx, 1, -1):
            if (x_clip[idx] - x_clip[idx-1]) == 0:
                continue
            slope = abs(y_clip[idx] - y_clip[idx-1])/(x_clip[idx] - x_clip[idx-1])
            if slope < eps:
                self.results['DC_loaded_V'] = (x_clip[idx], y_clip[idx])
                break
       
    def find_droops(self, tstart=None, tend=None, verbose=False):

        x_clip, y_clip = self._clip(tstart, tend)
        
        for prom in [100, 10, 1, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6, 1e-7, 1e-8, 1e-9, 1e-10]:
            valleys, _ = signal.find_peaks(1/y_clip, prominence=prom, width=2)
    
        #valleys, _ = signal.find_peaks(1/y_clip, prominence=1e-6)
        _ = self.mid_point(tstart, tend)
        min_idx = np.where(x_clip >= self.results['min'][0])[0][0]
        x_max = x_clip[np.where(y_clip == np.max(y_clip[min_idx:]))[0][0]]
        x_mid = (self.results['min'][0] + x_max)/2
        y_thrsh = y_clip[0]*0.99
        new_valleys = [valley_idx for valley_idx in valleys if x_clip[valley_idx] < x_mid and y_clip[valley_idx] < y_thrsh]

        try:
            self.results['min_1st_V'] = (x_clip[new_valleys[0]], y_clip[new_valleys[0]])
            self.results['min_2nd_V'] = (x_clip[new_valleys[1]], y_clip[new_valleys[1]])
            self.results['min_3rd_V'] = (x_clip[new_valleys[2]], y_clip[new_valleys[2]])
        except IndexError:
            pass
        
        self._find_DC_unloaded(tstart=tstart, tend=tend)
        for min, droop in zip(['min_1st_V', 'min_2nd_V', 'min_3rd_V'],
                              ['droop_1st_mV', 'droop_2nd_mV', 'droop_3rd_mV']):
            try:
                self.results[droop] = ('_', (self.results['DC_unloaded_V'][1] - self.results[min][1])*1e3)
            except KeyError:
                pass

        if verbose:
            for droop in ['min_1st_V', 'min_2nd_V', 'min_3rd_V',
                          'droop_1st_mV', 'droop_2nd_mV', 'droop_3rd_mV']:
                try:
                    print(f"{droop}:\t{self.results[droop]}")
                except KeyError:
                    pass

    def find_overshoots(self, tstart=None, tend=None, verbose=False):

        x_clip, y_clip = self._clip(tstart, tend)

        # Scale up data if it is low order to enable better peaks and valleys detection
        # y_clip = y_clip*1e6
        for prom in [100, 10, 1, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6, 1e-7, 1e-8, 1e-9, 1e-10]:
            peaks, _ = signal.find_peaks(y_clip*1e6, prominence=prom, width=2)
            if peaks.any():
                break
        
        
        peak_vals = [(peak_idx, x_clip[peak_idx], y_clip[peak_idx]) for peak_idx in peaks]
        peak_vals = sorted(peak_vals, key=itemgetter(1))[-3:]
        new_peaks = [peaks[0] for peaks in peak_vals]

        '''
        # Keep peaks that have times which are greater than the time of the maximum peak
        new_peaks = []
        for (peak_idx, x_val, y_val) in peak_vals:
            if x_val >= peak_vals[0][1]:
                new_peaks.append(peak_vals[0][0])
        '''
        
        # Remove false peaks
        _ = self.mid_point(tstart, tend)
        max_idx = np.where(x_clip >= self.results['max'][0])[0][0]
        x_min = x_clip[np.where(y_clip == np.min(y_clip[:max_idx]))[0][0]]
        x_mid = (x_min + self.results['max'][0])/2
        y_thrsh = self.results['min'][1]*1.01
        new_peaks = [peak_idx for peak_idx in new_peaks if x_clip[peak_idx] > x_mid and y_clip[peak_idx] > y_thrsh]

        try:
            self.results['max_1st_V'] = (x_clip[new_peaks[0]], y_clip[new_peaks[0]])
            self.results['max_2nd_V'] = (x_clip[new_peaks[1]], y_clip[new_peaks[1]])
            self.results['max_3rd_V'] = (x_clip[new_peaks[2]], y_clip[new_peaks[2]])
        except IndexError:
            pass

        self._find_DC_loaded(tstart=tstart, tend=tend)
        for max, over in zip(['max_1st_V', 'max_2nd_V', 'max_3rd_V'], ['over_1st_mV', 'over_2nd_mV', 'over_3rd_mV']):
            try:
                self.results[over] = ('_', (self.results[max][1] - self.results['DC_loaded_V'][1])*1e3)
            except KeyError:
                pass

        if verbose:
            for over in ['max_1st_V', 'max_2nd_V', 'max_3rd_V', 'over_1st_mV', 'over_2nd_mV', 'over_3rd_mV']:
                try:
                    if 'max' in over:
                        print(f"{over}:\t{self.results[over]}")
                    else:
                        print(f"{over.replace('over', 'overshoot')}:\t{self.results[over]}")
                except KeyError:
                    pass

    def droop_intervals(self):

        self.find_droops()
        self.find_overshoots()

        for droop in ['min_1st_V', 'min_2nd_V', 'min_3rd_V']:
            try:
                logger.info(f"{droop.replace('_', ' ')}:\t{(self.results[droop][0] - self.results['DC_unloaded_V'][0])*1e9:.3f} nsec")
            except KeyError:
                pass

        for over in ['max_1st_V', 'max_2nd_V', 'max_3rd_V']:
            try:
                logger.info(f"{over.replace('_', ' ')}:\t{(self.results[over][0] - self.results['DC_loaded_V'][0])*1e9:.3f} nsec")
            except KeyError:
                pass

    def resample(self, tsample, oversample=False, verbose=True):

        if not oversample:
            #new_x = np.linspace(self.data.x[0], self.data.x[-1], int((self.data.x[-1] - self.data.x[0])/tsample))
            new_x = np.arange(self.data.x[0], self.data.x[-1], tsample)
            inter_func = interp1d(self.data.x, self.data.y)
            new_y = inter_func(new_x)
        else:
            new_x = np.array([])
            num_samples = np.floor((self.data.x[1:] - self.data.x[:-1])/tsample)
            for idx, nsamples in enumerate(num_samples):
                if nsamples > 0:
                    #new_x = np.append(new_x, np.linspace(self.data.x[idx], self.data.x[idx+1], int(nsamples), False))
                    new_x = np.arange(self.data.x[0], self.data.x[-1], tsample)
                else:
                    new_x = np.append(new_x, self.data.x[idx])
            new_x = np.append(new_x, self.data.x[-1])

            inter_func = interp1d(self.data.x, self.data.y)
            new_y = inter_func(new_x)

        history = self.history + [f'Resample: sample interval = {tsample} sec | elements = {len(new_y)} | Oversample = {oversample}']
        prim.Wave.wave_num += 1
        resamp_wave = prim.Wave(DataVector(x=new_x, y=new_y,
                                      x_unit=self.x_unit,
                                      y_unit=self.y_unit,
                                      wave_name=f'{self.wave_name}_resampled{prim.Wave.wave_num-1}',
                                      file_name=self.file_name,
                                      path = self.path,
                                      proc_hist=history
                                     )
                          )

        if verbose:
            logger.info(f'Sampling interval: {tsample*1e9} nsec\nBefore resmapling: '
                  f'{len(self.data.y)} points, after resmapling: {len(new_y)} points')
            util = ld.Waveforms()
            util.path = resamp_wave.path
            util.plot_overlay([self, resamp_wave])
        return resamp_wave

    def filt(self, f, tstart=None, tend=None):

        resampled_data = self.clip(tstart, tend).resample(1/f.fs, oversample=False, verbose=False)
        new_x, new_y = resampled_data.data.x, resampled_data.data.y
        DC_comp = np.mean(new_y)
        new_y = new_y - DC_comp
        if not f.keep_DC and not f.filter_type == 'lowpass':
            DC_comp = 0
        
        proc_hist1 = f'Filter: {f.filter_name} {f.filter_type} | Order: {f.order} | Ripple: {f.ripple} dB | Attanuation: {f.att} dB | '
        proc_hist2 = f'Lowpass cutoff freq.: {f.cutoff_freq[0]} Hz | Highpass cutoff freq.: {f.cutoff_freq[1]} Hz' if isinstance(f.cutoff_freq, tuple) else f'{f.filter_type.capitalize()} cutoff freq.: {f.cutoff_freq} Hz'
        history = self.history + [proc_hist1, proc_hist2]

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=new_x, y=DC_comp + signal.sosfilt(f.filter_func(), new_y),
                               x_unit=self.x_unit,
                               y_unit=self.y_unit,
                               wave_name=f'{self.wave_name}_filtered{prim.Wave.wave_num-1}',
                               file_name=self.file_name,
                               path=self.path,
                               proc_hist=history
                              )
                    )

    def _save_csv(self, sep, fname, subckt_name=None, decimals=None, header=True):

        new_x = self.data.x if decimals is None else np.around(self.data.x, decimals)
        wave = pd.DataFrame({'time': new_x, 'amp': self.data.y})
        if header:
            header = ['t', f'{self.wave_name}' if subckt_name is None else subckt_name]
        else:
            header = None
        wave.to_csv(f'{fname}.csv', sep=sep, index=False, header=header)

    def _save_PWL(self, fname, subckt_name=None):

        subckt_name = f"Icct_{self.wave_name.replace('.', '_').replace('(', '').replace(')', '')}" \
                        if subckt_name is None else subckt_name
        with open(f'{fname}.inc', 'wt') as f:
            f.write(f".subckt\t{subckt_name}\tp1\tp2\tdelay=0\n\n")
            f.write('+ PWL (\n')
            for time, amp in zip(self.data.x, self.data.y):
                f.write(f'+ {time}\t{amp}\n')
            f.write('+ TD=delay\n+ R)\n.ends\n')

    def _save_gpoly(self, vnom, fname, subckt_name=None):

        subckt_name = f"Icct_{self.wave_name.replace('.', '_').replace('(', '').replace(')', '')}" \
                        if subckt_name is None else subckt_name
        with open(f'{fname}.inc', 'wt') as f:
            f.write(f".subckt\t{subckt_name}\tvdd\tsubcktgnd\tdelay=0\tvnom={vnom}\n\n")
            f.write(f"Gpoly_icct\tvdd subcktgnd poly(2) controlnode "
                    f"0 vdd subcktgnd 0 0 0 0 '1/vnom'\n")
            f.write(f'.probe tran i(Gpoly_icct)\n\n')
            f.write(f'Vdd\tcontrolnode\tsubcktgnd\n+ PWL(\n')
            for time, amp in zip(self.data.x, self.data.y):
                f.write(f'+ {time}\t{amp}\n')
            f.write('+ TD=delay\n+ R)\n.ends\n')
        
    def save(self, format_type, fname=None, vnom=1, sep=',',
                subckt_name=None, suffix=None, decimals=None, header=True):

        save_func = {'csv': self._save_csv, 'PWL': self._save_PWL, 'gpoly': self._save_gpoly}
        suffix = '' if suffix is None else '_' + suffix

        fname = self.wave_name if fname is None else fname
        fname = fname.replace('/', '').replace('\\', '').replace('(', '').replace(')', '')
        fname = Path(self.path) / Path('new_waveforms') / Path(f'{fname}{suffix}')

        if format_type == 'csv':
            save_func[format_type](sep, fname, subckt_name, decimals, header)
        elif format_type == 'gpoly':
            save_func[format_type](vnom, fname, subckt_name)
        elif format_type == 'PWL':
            save_func[format_type](fname, subckt_name)

    def shift(self, tshift=None):

       # If shift=None shift to zero
       tshift = -self.data.x[0] if tshift is None else tshift
       
       prim.Wave.wave_num += 1
       return prim.Wave(DataVector(x=self.data.x+tshift, y=self.data.y,
                              x_unit=self.x_unit,
                              y_unit=self.y_unit,
                              wave_name=f'{self.wave_name}_shifted{prim.Wave.wave_num-1}',
                              file_name=self.file_name,
                              path=self.path,
                              proc_hist=self.history + [f'Shift: shift time = {tshift} sec']
                             )
                  )
    
    def resonance_tuning(self, low2high_start, high2low_start,
                         freq, duty=50):
        
        thigh = (1/freq)*(duty/100)
        tlow = (1/freq) - thigh

        if thigh > high2low_start:
            logger.info(f'Cannot tune {self.wave_name} at {freq*1e-6:.3f} MHz. Increase frequency.')
            return
        
        high = self.clip(low2high_start, low2high_start + thigh)
        low = self.clip(high2low_start, high2low_start + tlow)

        return high.snap(low).shift()

    def create_BIB_DDR(self, preamble_start, preamble_end, postamble_start, postamble_end,
                             freq, duty=50, num_bits=1, bits_per_packet=8):
        
        preamble = self.clip(preamble_start, preamble_end)
        data = self.clip(preamble_end, postamble_start)
        postamble = self.clip(postamble_start, postamble_end)

        tpre = preamble_end - preamble_start
        tdata = postamble_start - preamble_end
        tpost = postamble_end - postamble_start

        tburst = (1/freq)*(duty/100)
        tburst = tburst - tpre
        tbit = tdata/num_bits
        total_bits = int(round((tburst / tbit) / bits_per_packet)*bits_per_packet)
        new_tburst = tpre + total_bits*tbit
        new_tidle = 1/freq - new_tburst
        new_duty = new_tburst/(1/freq)*100

        if total_bits == 0:
            logger.warning(f'Frequency {freq} [Hz] is too high. Cannot create BIB pattern.')
            return

        data_portion = data.create_BIB(freq=1/((total_bits - (bits_per_packet - num_bits))*tbit), duty=100)
        #data_portion = data_portion.clip(tend=data_portion.data.x[-3]).shift(preamble_end)
        data_portion = data_portion.shift(preamble_end)
        
        if tpost >= new_tidle:
            idle_portion = postamble.shift().shift(data_portion.data.x[-1]).clip(tend=1/freq)
        else:
            idle_portion = postamble.shift().shift(data_portion.data.x[-1]) # Bring to zero and then shift
            idle_portion.data.y[-1] = preamble.data.y[0]
            idle_portion.data.x[-1] = new_tburst + new_tidle

        # Stitching everything together
        data_portion.data.y[0] = 0
        data_portion.data.x[0] = preamble.data.x[-1]
        data_portion.data.y[-1] = 0
        data_portion.data.x[-1] = idle_portion.data.x[0]
    
        logger.info(f'Duration per bit = {tbit:.4} [sec], Total number of bits = {total_bits}')
        logger.info(f'BIB burst time = {new_tburst:.4} [sec], BIB duty cycle = {new_duty:.3}%')
        logger.info(f'Average bit = {data.average():.4} [A]')

        return preamble + data_portion + idle_portion

    def extend(self, ext_time, preamble_start=None, preamble_end=None,
               data_start=None, data_end=None, postamble_start=None,
               postamble_end=None):

        return self.create_BIB(freq=1/ext_time, duty=100, preamble_start=preamble_start,
                               preamble_end=preamble_end, data_start=data_start,
                               data_end=data_end, postamble_start=postamble_start,
                               postamble_end=postamble_end)

    def denoise(self, noise_att=0.5, perc_reduc=50, verbose=True):

        if noise_att >= 0 and noise_att <= 1:
            noise_att = int(1000*noise_att)
            if noise_att % 2 == 0:
                noise_att += 1
        else:
            noise_att = 1001

        y = median_filter(self.data.y, noise_att)
        x = self.data.x
        diff_y = np.abs(y[1:] - y[:-1])

        for reduce in np.linspace(0, np.max(diff_y), 1000):
            new_y = y[:-1][diff_y > reduce]
            new_x = x[:-1][diff_y > reduce]
            if ((1 - len(new_y)/len(y))*100) > perc_reduc:
                break
        
        prim.Wave.wave_num += 1
        denoised_wave = prim.Wave(DataVector(x=new_x, y=new_y,
                                x_unit=self.x_unit,
                                y_unit=self.y_unit,
                                wave_name=f'{self.wave_name}_denoised{prim.Wave.wave_num-1}',
                                file_name=self.file_name,
                                path=self.path,
                                proc_hist=self.history + [f'Denoised with {noise_att} noise level']
                                )
                    )

        if verbose:
            print(f'Before denoising: '
                    f'{len(self.data.y)} points, after denoising: '
                    f'{len(denoised_wave.data.y)} points')
            print(f'noise_att = {noise_att}, % reduction = {(1 - len(new_y)/len(y))*100:.2f}%')
            util = ld.Waveforms()
            util.path = self.path
            util.plot_overlay([self, denoised_wave])
        
        return denoised_wave

    def create_BIB(self, freq, duty=50, idle=None, preamble_start=None, preamble_end=None,
                         data_start=None, data_end=None, postamble_start=None, postamble_end=None):

        pre, post = None, None
        if preamble_end is None:
            tpre = 0
        else:
            if preamble_start is None:
                preamble_start = self.data.x[0]
            tpre = preamble_end - preamble_start
            pre = self.clip(preamble_start, preamble_end).shift()
            pre.data.y[-1] = 0
        if postamble_start is None:
            tpost = 0
        else:
            if postamble_end is None:
                postamble_end = self.data.x[-1]
            tpost = postamble_end - postamble_start
            post = self.clip(postamble_start, postamble_end).shift()

        wave = self.clip(data_start, data_end).shift()
        tburst = (duty/100)*(1/freq) - (tpre + tpost)
        tidle = (1 - duty/100)*(1/freq)

        new_wave = deepcopy(wave)
        
        tcycle = new_wave.data.x[-1] - new_wave.data.x[0]
        num_cycles = tburst/tcycle
        if num_cycles >= 1:
            for num_cycle in range(1, int(np.ceil(num_cycles))):
                shifted_wave = wave.shift(num_cycle*tcycle)
                new_wave.data.y[-1] = 0
                new_wave.data.x[-1] = shifted_wave.data.x[0]
                new_wave = new_wave + shifted_wave

            new_wave = new_wave.clip(tend=tburst)
            if pre is None:
                final_wave = new_wave
            else:
                pre.data.y[-1] = 0
                final_wave = pre + new_wave.shift(pre.data.x[-1])
            if post is not None:
                final_wave.data.y[-1] = 0
                final_wave = final_wave + post.shift(final_wave.data.x[-1])
            
            if duty < 100:
                if idle is None:
                    idle = final_wave.data.y[0]
                elif isinstance(idle, tuple):
                    idle = self.average(idle[0], idle[1])
                final_wave.data.y[-2] = idle
                final_wave.data.y[-1] = idle
                final_wave.data.x[-1] = final_wave.data.x[-1] + tidle
            
            prim.Wave.wave_num += 1
            final_wave.wave_name = f'{final_wave.wave_name}_BIB{prim.Wave.wave_num-1}'
            final_wave.history = [f'BIB: freq={freq*1e-6} MHz, duty cycle={duty}%, '
                                  f'idle={idle} {final_wave.y_unit}']
            return final_wave
        else:
            logger.info(f'\nCannot create BIB pattern.\n'
                    f'Burst time {tburst*1e9:.3f} nsec is smaller '
                    f'than the preamble ({tpre*1e9:.3f} nsec) '
                    f'+ data ({tcycle*1e9:.3f} nsec) '
                    f'+ postamble ({tpost*1e9:.3f} nsec).')

    def snap(self, wave):

        y_value = self.data.y[-1]
        self.data.y[-1] = 0
        snapped_wave = self + wave.shift(self.data.x[-1] - wave.data.x[0])
        self.data.y[-1] = y_value
        
        return snapped_wave

    def fft(self, n=None, win_type=None, verbose=True):

        wins = {None: lambda n: np.ones(n), 'blackman': blackman, 'hann': hann,
                    'hamming': hamming, 'bartlett': bartlett}
        
        # Find the next power of 2 to icrease FFT algorithm speed
        # n = int(2**np.ceil((np.log10(len(self.data.x))/np.log10(2)))) if n is None else n
        n = int(len(self.data.x)) if n is None else n

        ts = (self.data.x[-1] - self.data.x[0])/len(self.data.x)
        fs = 1/ts
        new_wave = self.resample(ts, verbose=False)

        y_fft = fft(new_wave.data.y*wins[win_type](n), n)
        f = np.arange(0, len(y_fft))*fs/len(y_fft)
        #f = fftfreq(n, ts)
        new_y = (2/n)*np.abs(y_fft)
        new_y[0] = 0.5*new_y[0]

        prim.Wave.wave_num += 1
        fft_wave = prim.Wave(DataVector(x=f[:int(np.ceil(n/2))], y=new_y[:int(np.ceil(n/2))],
                                   x_unit='Hz',
                                   y_unit='',
                                   wave_name=f'{self.wave_name}_fft_single_side{prim.Wave.wave_num-1}',
                                   file_name=self.file_name,
                                   path=self.path,
                                   proc_hist=self.history + [f'FFT with {ts*1e9:.3f} nsec sampling time, single sided']
                                  )
                        )
        complex_wave = prim.Wave(DataVector(x=f, y=y_fft,
                                        x_unit='Hz',
                                        y_unit='',
                                        wave_name=f'{self.wave_name}_fft_double_side{prim.Wave.wave_num-1}',
                                        file_name=self.file_name,
                                        path=self.path,
                                        proc_hist=self.history + [f'FFT with {ts*1e9:.3f} nsec sampling time, double sided']
                                    )
                            )

        if verbose:
            logger.info(f'FFT({self.wave_name}) with {ts*1e9:.3f} nsec sampling time')
            fft_wave.plot(x_scale='M')
        
        return fft_wave, complex_wave

    def ifft(self, n=None, verbose=True):

        n = int(len(self.data.x)) if n is None else n
        y_ifft = ifft(self.data.y, n)
        ts = 1/(self.data.x[-1] - self.data.x[0])
        t = np.arange(0, len(y_ifft))*ts

        ifft_wave = prim.Wave(DataVector(x=t, y=np.real(y_ifft),
                                   x_unit='sec',
                                   y_unit='',
                                   wave_name=f'{self.wave_name}_ifft{prim.Wave.wave_num-1}',
                                   file_name=self.file_name,
                                   path=self.path,
                                   proc_hist=self.history + [f'IFFT with {ts*1e9:.3f} nsec sampling time']
                                  )
                        )

        if verbose:
            logger.info(f'IFFT({self.wave_name}) with {ts*1e9:.3f} nsec sampling time')
            ifft_wave.plot(x_scale='M')
        
        return ifft_wave

    def pow(self, n):

        if n == 0:
            return self/self
        if n < 0:
            result = 1/self
            n = -1*n
        else:
            result = self

        for _ in range(n-1):
            result *= self
        
        return result

    def mean_square_error(self, other, tstart=None, tend=None):

        self = self.clip(tstart, tend)
        other = other.clip(tstart, tend)
        mse = ((self - other).pow(2)).average()
    
        return mse

    def detect_settle(self):
        '''Finds a settling time window of the waveform.
        Settling is assumed to accure at the end of the waveform.

        :return: Settle time window
        :rtype: tuple
        '''

        new_wave = -1*self
        try:
            time_windows = new_wave.detect_change_points(plot_fig=False)
        except IndexError:
            time_windows = self.detect_change_points(plot_fig=False)

        self.results['settle'] = time_windows[-1]
        return self.results['settle']

    def detect_change_points(self, plot_fig=True, slew_perc=None):
        '''Detects change points in the x-axis domain.
        This algorithm assumes one cycle of the periodic waveform
        as well as up to three distinguished regions.

        :param plot_fig: If True indiciating to generate a plot with
        a marked time window, defaults to True
        :type plot_fig: bool, optional
        :return: Table with the change points and the corresponding statistics
        :rtype: list or pandas DataFrame, depending on the value of plot_fig
        '''

        change_points = []
        denoised_wave = self.denoise(0.1, 0, verbose=False)
        # denoised_wave.plot()

        # Detect change points in the middle section
        max_thrsh = denoised_wave.maximum()

        #max_thrsh = self.maximum()
        middle_range = self.data.x[self.data.y > max_thrsh*0.97]
        change_points.append((middle_range[0], middle_range[-1]))
        
        # Detect change points in the left section
        if np.abs((middle_range[1] - self.data.x[1])/self.data.x[1]) > 0.05:
            left_wave = self.clip(tend=(self.data.x[0] + self.data.x[-1])/2)
            min_thrsh = self.average(tend=self.data.x[int(0.1*len(self.data.x))])
            left_range = left_wave.data.x[left_wave.data.y <= min_thrsh]
            
            if left_range[-1] > middle_range[0]:
                change_points = [(left_range[0], middle_range[0]*0.9)] + change_points
            else:
                change_points = [(left_range[0], left_range[-1])] + change_points

        # Detect change points in the right section
        if np.abs((middle_range[-1] - self.data.x[-1])/self.data.x[-1]) > 0.05:
            right_wave = self.clip(tstart=(self.data.x[0] + self.data.x[-1])/2)
            min_thrsh = self.average(tstart=self.data.x[int(0.9*len(self.data.x))])
            right_range = right_wave.data.x[right_wave.data.y <= min_thrsh]

            if right_range[0] < middle_range[-1]:
                change_points.append((middle_range[-1]*1.05, right_range[-1]))
            else:
                change_points.append((right_range[0], right_range[-1]))

        change_points = sorted(change_points, key=itemgetter(0))

        # Clean up change points
        change_points = [(left, right) for (left, right) in change_points if left < right]
        #print(change_points)

        if plot_fig:
            if slew_perc is None:
                result = self.stat_ranges(*change_points)
            else:
                result = (self.stat_ranges(*change_points),
                            self.detect_slew_rate(denoised_wave=denoised_wave,
                                                    slew_perc=slew_perc)
                    )
        else:
            if slew_perc is None:
                result = change_points
            else:
                result = (change_points
                            + [self.detect_slew_rate(denoised_wave=denoised_wave,
                                                    slew_perc=slew_perc)]
                    )

        return result

    def stat_ranges(self, *xaxis_points):
        '''Creates table of the change points and the corresponding statistics.
        It also saves figures with the corresponding waveform and regions.

        :param xaxis_points: Start and end points of the region
        :type xaxis_points: tuple
        :return: Table with the change points and the corresponding statistics
        :rtype: pandas DataFrame
        '''
        
        plt.plot(self.data.x*1e9, self.data.y*1e3)
        stats = {'Start Time [nsec]': [], 'End Time [nsec]': [],
                    'Min. [mA]': [], 'Max. [mA]': [], 'Avg. [mA]': []}
        ycenter = (self.maximum() + self.minimum())/2
        for xstart, xend in xaxis_points:
            stats['Start Time [nsec]'].append(f'{xstart*1e9:.3f}')
            stats['End Time [nsec]'].append(f'{xend*1e9:.3f}')
            clipped_wave = self.clip(xstart, xend)
            stats['Min. [mA]'].append(f'{clipped_wave.minimum()*1e3:.3f}')
            stats['Max. [mA]'].append(f'{clipped_wave.maximum()*1e3:.3f}')
            stats['Avg. [mA]'].append(f'{clipped_wave.average()*1e3:.3f}')
            plt.axvspan(xstart*1e9, xend*1e9, facecolor='g', alpha=0.2)
            plt.text(xstart*1e9, ycenter*1e3,
                    f"Avg: {stats['Avg. [mA]'][-1]} [mA]\n"
                    f"Max: {stats['Max. [mA]'][-1]} [mA]\n"
                    f"Min: {stats['Min. [mA]'][-1]} [mA]",
                    fontsize=8)
        
        plt.grid()
        plt.title(f'{self.file_name}\n{self.wave_name}')
        plt.xlabel('Time [nsec]')
        plt.ylabel('Amplitude [mA]')
        plt.savefig(os.path.join(self.path, 'plots', f'{self.wave_name}.png'),
                    bbox_inches='tight')
        plt.close()
        return pd.DataFrame(stats)

    def detect_slew_rate(self, denoised_wave=None, slew_perc=90):
        '''Automatically detects slew rate (A/sec) of a waveform.

        :param denoised_wave: Noise reduced waveformed.
        If not given, a noise reduced waveform will be generated, defaults to None.
        :type denoised_wave: :class:`primitives.Wave()`, optional
        :param slew_perc: Percent of the amplitude to use
        when calculating the slew rate. For example, 90 means 90%-10%,
        80 means 80%-20%, defaults to 90 percent.
        :type slew_perc: int, optional
        :return: Slew rate, start time, and end time
        :rtype: tuple[float. float, float]
        '''
    
        if denoised_wave is None:
            denoised_wave = self.denoise(0.9, 0, verbose=False)
        
        max_point = denoised_wave.maximum()
        mid_point = denoised_wave.mid_point()
        min_point = denoised_wave.minimum()
        tot_range = (max_point - min_point)*((2*slew_perc - 100)/100)

        # Compare the x values of the maximum and minimum points
        if denoised_wave.results['max'][0] < denoised_wave.results['min'][0]:
            ystart = mid_point + tot_range/2
            yend = mid_point - tot_range/2
        else:
            ystart = mid_point - tot_range/2
            yend = mid_point + tot_range/2

        tstart = self.x_at_y(ystart)
        tend = self.x_at_y(yend)

        return self.slew_rate(tstart, tend), tstart, tend

    def slew_rate(self, tstart=None, tend=None):
        '''Calculates slew rate (amp/sec) of a waveform between tstart and tend.

        :param tstart: Start time of the truncation window in seconds, defaults to None
        :type tstart: float, None, optional
        :param tend: End time of the truncation window in seconds, defaults to None
        :type tend: float, None, optional
        :return: Slew rate of the waveform
        :rtype: float
        '''

        try:
            x_clip, y_clip = self._clip(tstart, tend)
            self.results['slew'] = (y_clip[-1] - y_clip[0])/(x_clip[-1] - x_clip[0])
        except IndexError:
            return None
        
        return self.results['slew']

class GroupOps():

    def __init__(self):

        self.port_loc = {}
        self.name_conv = {}
        self.sheets = []
        self.plt = Plotter()

    def _waves_to_plot(self, wave_names):

        names = self.wave_names(False) if wave_names is None else wave_names
        names = [names] if not isinstance(names, list) else names
        waves_to_plot = []
        # Check if the list contain Wave objects.
        # If so, extract wave_name and insert into the waveform pool
        for name in names:
            if isinstance(name, (prim.Wave, prim.WaveNet)):
                # name here is a Wave or prim.WaveNet object
                waves_to_plot.append(name)
            elif isinstance(name, ld.Waveforms):
                waves_to_plot += name.waves.values()
            else:
                waves_to_plot.append(self.waves[name])
        
        return waves_to_plot

    def plot_stack(self, wave_names=None, x_scale='u', y_scale='', xaxis_type='linear',
                   yaxis_type='linear', clr=None, out_file=None):

        out_file = 'stack_waves' if out_file is None else out_file
        output_file(os.path.join(self.path, 'plots', f'{out_file}.html'))
        all_plots = []
        
        clr_cycle = cycle(Category20[20]) if clr is None else cycle([clr])
        for wave in self._waves_to_plot(wave_names):
            all_plots.append(self.plt.single_plot(wave, x_scale=x_scale,
                                                  y_scale=y_scale, clr=next(clr_cycle),
                                                  xaxis_type=xaxis_type,
                                                  yaxis_type=yaxis_type
                                                 )
                            )

        show(column(all_plots))

    def cplot_stack(self, wave_names=None, x_scale='M', y_scale='', xaxis_type='log',
                    clr=None, net_type='z', i=None, j=None, out_file=None):

        net_type = net_type.lower()
        out_file = 'stack_waves' if out_file is None else out_file
        output_file(os.path.join(self.path, 'plots', f'{out_file}.html'))
        waves = self._waves_to_plot(wave_names)

        # Select only prim.WaveNet objects
        for wave in waves.copy():
            if not isinstance(wave, prim.WaveNet):
                waves.remove(wave)

        all_plots = []
        clr_cycle = cycle(Category20[20]) if clr is None else cycle([clr])
        
        if not i is None or not j is None:
            waves[0].plot(net_type=net_type, x_scale=x_scale,
                            y_scale=y_scale, xaxis_type=xaxis_type,
                            clr=clr, i=i, j=j)
            return

        for wave in waves:
            i, j = wave.port_num, wave.port_num
            clr = next(clr_cycle)

            for net_attr, cwaves in wave.net_attr(net_type.lower(), i, j).items():
                all_plots.append(
                        self.plt.single_plot(cwaves[0], x_scale=x_scale, y_scale=y_scale,
                                    xaxis_type=xaxis_type, yaxis_type=cwaves[1],
                                    clr=clr,
                                    suffix=f'{wave.wave_name} {net_attr.capitalize()} '
                                            f'{net_type.upper()}({i}, {j})',
                                    y_title=f'{net_attr.capitalize()} {net_type.upper()}',
                                    stretch_mode=False
                                    )
                            )

        show(column(all_plots))

    def plot_overlay(self, wave_names=None, x_scale='u', y_scale='', xaxis_type='linear',
                     yaxis_type='linear', out_file=None, clr=None):

        out_file = 'overlay_waves' if out_file is None else out_file
        output_file(os.path.join(self.path, 'plots', f'{out_file}.html'))

        if isinstance(wave_names, list) and isinstance(wave_names[0], list): # Check if it's a list of lists
            plots = []
            for waves in wave_names:
                plots.append(self.plt.plot_overlay(self._waves_to_plot(waves),
                                   x_scale=x_scale, y_scale=y_scale,
                                   xaxis_type=xaxis_type, clr=clr,
                                   yaxis_type=yaxis_type,
                                   stretch_mode=None
                                  ))
            show(column(plots))
        else:
            show(self.plt.plot_overlay(self._waves_to_plot(wave_names),
                                       x_scale=x_scale, y_scale=y_scale,
                                       xaxis_type=xaxis_type, clr=clr,
                                       yaxis_type=yaxis_type
                                      )
                )

    def cplot_overlay(self, wave_names=None, x_scale='M', y_scale='', xaxis_type='log',
                        clr=None, net_type='z', i=None, j=None, out_file=None):

        out_file = 'overlay_waves' if out_file is None else out_file
        output_file(os.path.join(self.path, 'plots', f'{out_file}.html'))
        waves = self._waves_to_plot(wave_names)

        # Select only prim.WaveNet objects
        for wave in waves.copy():
            if not isinstance(wave, prim.WaveNet):
                waves.remove(wave)

        if i is not None or j is not None:
            waves[0].plot(net_type=net_type, x_scale=x_scale,
                            y_scale=y_scale, xaxis_type=xaxis_type,
                            clr=clr, i=i, j=j)
            return

        all_plots = defaultdict(list)
        plot_yaxis_type = {}
        for wave in waves:
            i, j = wave.port_num, wave.port_num
            for net_attr, cwaves in wave.net_attr(
                                    net_type=net_type.lower(), i=i, j=j).items():
                all_plots[net_attr].append(cwaves[0])
                plot_yaxis_type[net_attr] = cwaves[1]

        plots_to_stack = []
        for attr, plots in all_plots.items():
            plots_to_stack.append(self.plt.plot_overlay(plots, title=f'{net_type.upper()} Parameters',
                                    yaxis_title=f'{attr.capitalize()} {net_type.upper()}',
                                    x_scale=x_scale, y_scale=y_scale, stretch_mode=None, clr=clr,
                                    xaxis_type=xaxis_type, yaxis_type=plot_yaxis_type[attr]
                                                        )
                                )

        show(column(plots_to_stack))

    def port_selection(self, *ports):

        wave_names = list(self.waves.keys())
        selected_ports = []
        for port in ports:
            port = str(port)
            if port.isnumeric():
                selected_ports.append(wave_names[int(port)])
            elif ':' in port:
                start_idx, end_idx = port.split(':')
                start_idx = 0 if start_idx == '' else int(start_idx)
                end_idx = None if end_idx == '' else int(end_idx)
                selected_ports += wave_names[start_idx:end_idx]
            else:
                selected_ports.append(port)

        return selected_ports

    def _terminate(self, *terms):
        
        conns = {'open': [], 'short': [], 'circuit': [], 'port': []}
        for term_type, ports in terms:
            if term_type == 'circuit':
                circuit = ports
                term_port_names = self.port_selection(*circuit.port_nums)
            else:
                term_port_names = self.port_selection(*ports)
            conns[term_type] += [self.waves[name].port_num for name in term_port_names]

        all_term_ports = conns['open'] + conns['short'] + conns['circuit']
        conns['port'] = [wave.port_num for wave in self.waves.values()
                            if wave.port_num not in all_term_ports
                        ]

        term_net = self.get_net_param()[1]
        if term_net.f[0] == 0:
            term_net.f[0] = 0.01
            
        freq = rf.Frequency.from_f(term_net.f, unit='Hz')
        z0 = term_net.z0[0, 0]
        s_open = np.full((freq.f.shape[0], 1, 1), 1.0 + 0.0j, dtype=complex)
        s_short = np.full((freq.f.shape[0], 1, 1), -1.0 + 0.0j, dtype=complex)
        
        connections = []
        # Connect terminations
        for port_num in conns['short']:
            connections.append([(term_net, port_num),
                                    (Network(frequency=freq, s=s_short, name=f'short{port_num}', z0=z0), 0)]
                                )
        for port_num in conns['open']:
            connections.append([(term_net, port_num),
                                    (Network(frequency=freq, s=s_open, name=f'open{port_num}', z0=z0), 0)]
                                )
        for port_num in conns['circuit']:
            connections.append([(term_net, port_num), (circuit.cir_map[port_num], 0)])

        # Connect ports
        for port_num in conns['port']:
            connections.append([(term_net, port_num), (rf.Circuit.Port(freq, f'port{port_num}', z0=z0), 0)])

        new_net = rf.Circuit(connections).network

        return conns, new_net

    def terminate(self, *terms):

        conns, term_net = self._terminate(*terms)
        term_net.name = self.get_net_param()[1].name
        

        new_waves = []
        port_idx = 0
        for wave in self.waves.values():
            if wave.port_num in conns['port']:
                new_waves.append(prim.WaveNet(DataVector(x=term_net.f,
                                                        y=term_net.z_mag[:, port_idx, port_idx],
                                                        x_unit=wave.x_unit,
                                                        y_unit=wave.y_unit,
                                                        wave_name=wave.wave_name,
                                                        file_name=wave.file_name,
                                                        path=wave.path,
                                                        proc_hist=wave.history
                                                        ),
                                                term_net,
                                                wave.ts_info,
                                                port_idx
                                            )
                                )
                port_idx += 1

        return self.group(*new_waves)

    def _auto_term(self, port_name, other_ports=False, term_type='short'):

        term_idx = []
        logger.info(f'Ports to be terminated {term_type}:')
        for name, port in self.waves.items():
            if (port_name.lower() in name.lower() and not other_ports) \
                        or (not port_name.lower() in name.lower() and other_ports):
                term_idx.append(port.port_num)
                logger.info(f'\t{name}')

        return self.terminate((term_type, term_idx))

    def get_net_param(self, freq=None):

        net = list(self.waves.values())[0].net
        if freq is not None:
            try:
                freq_idx = np.where(freq == net.f)[0][0]
            except IndexError: # The required frequency does not exist and interpolation is required
                logger.warning(f'Frequency {freq*1e-6} MHz is not defined in the touchstone file.\n'
                        f'Interpolating S-parameters at {freq*1e-6:.2f} MHz.')
                new_f = np.append(net.f, freq)
                new_f.sort()
                new_f = rf.Frequency.from_f(new_f, unit='Hz')
                net = net.interpolate_from_f(new_f)
                freq_idx = np.where(freq == net.f)[0][0]
        else:
            freq_idx = None

        return (freq_idx, net)

    def ind_char(self, freq, port_name='sw', net=None):

        if net is None:
            # Find all non SW ports in order to short them
            group = self._auto_term(port_name, other_ports=True)
        idx_freq, net = group.get_net_param(freq)
        
        L11 = np.mean(np.diag(net.z_im[idx_freq])/(2*np.pi*freq))
        L21 = []
        for col in range(net.number_of_ports):
            L21.append(np.sort(net.z_im[idx_freq][:, col])[-2])
        L21 = np.mean(L21)/(2*np.pi*freq)
        K = L21/L11
        Rac = np.mean(np.diag(net.z_re[idx_freq]))
        Rdc = np.mean(np.diag(net.z_re[0]))
        Q = (2*np.pi*freq*L11)/Rac
        Ltrans = 1/(np.linalg.multi_dot(
                                        [np.ones((1, net.number_of_ports)),
                                        np.linalg.inv(net.z_im[idx_freq]/(2*np.pi*freq)),
                                        np.ones((net.number_of_ports, 1))]
                                        )
                    )[0][0]

        L_mat = pd.DataFrame(net.z_im[idx_freq]/(2*np.pi*freq)*1e9)
        R_mat = pd.DataFrame(net.z_re[idx_freq]*1e3)
        
        logger.info(f'\nL11\t{L11*1e9:.4f} nH\nL21\t{L21*1e9:.4f} nH\n'
                f'K\t{K:.4f}\nRac\t{Rac*1e3:.4f} mOhm\nRdc\t{Rdc*1e3:.4f} mOhm\n'
                f'Q\t{Q:.4f}\nLtrans\t{Ltrans*1e9:.4f} nH')

        logger.info(f'\nInductance matrix (nH) at {freq*1e-6} MHz')
        logger.info(L_mat)

        logger.info(f'\nResistance (mOhm) matrix at {freq*1e-6} MHz')
        logger.info(R_mat)

    def renormalize_imp(self, refz=50):

        # Any port will change the ref. impednace of all ports
        self.waves[self.wave_names(verbose=False)[0]].net.renormalize(refz)

    def save_touchstone(self, fname=None, refz=None):

        fname = os.path.join(self.path, self.waves[0].net.name.split('.')[0]) \
                if fname is None else os.path.join(self.path, fname.split('.')[0])

        # Using any port will save all the s-parameters
        self.waves[self.wave_names(verbose=False)[0]].net.write_touchstone(
                        filename=fname, r_ref=refz
                    )

    def sparam_props(self):

        for wave in self.waves.values():
            wave.sparam_props()

    def _time_vector(self, tstart, tend):

        if not isinstance(tstart, list):
            tstart = [tstart]*len(self.waves)
        if not isinstance(tend, list):
            tend = [tend]*len(self.waves)

        return tstart, tend

    def maximum(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.maximum(tstart, tend)
                for (wave_name, wave), tstart, tend in zip(self.waves.items(), ts, te)}

    def minimum(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.minimum(tstart, tend)
                for (wave_name, wave) ,tstart, tend in zip(self.waves.items(), ts, te)}
        
    def pk2pk(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.pk2pk(tstart, tend)
                for (wave_name, wave) ,tstart, tend in zip(self.waves.items(), ts, te)}

    def mid_point(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.mid_point(tstart, tend)
                for (wave_name, wave) ,tstart, tend in zip(self.waves.items(), ts, te)}

    def average(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.average(tstart, tend)
                for (wave_name, wave) ,tstart, tend in zip(self.waves.items(), ts, te)}

    def find_droops(self, tstart=None, tend=None, verbose=False):

        all_results = {}
        for wave_name, wave in self.waves.items():
            if verbose:
                print(wave_name)
                wave.find_droops(tstart, tend, verbose)
            else:
                result = {}
                wave.find_droops(tstart, tend, verbose)
                for droop in ['min_1st_V', 'min_2nd_V', 'min_3rd_V',
                                'droop_1st_mV', 'droop_2nd_mV', 'droop_3rd_mV',
                                'DC_unloaded_V']:
                    try:
                        result[droop] = wave.results[droop]
                    except KeyError:
                        pass
                if result:
                    all_results[wave_name] = result
            
        if not verbose:
            return all_results

    def find_overshoots(self, tstart=None, tend=None, verbose=False):

        all_results = {}
        for wave_name, wave in self.waves.items():
            if verbose:
                print(wave_name)
                wave.find_overshoots(tstart, tend, verbose)
            else:
                result = {}
                wave.find_overshoots(tstart, tend, verbose)
                for over in ['max_1st_V', 'max_2nd_V', 'max_3rd_V',
                                'over_1st_mV', 'over_2nd_mV', 'over_3rd_mV',
                                'DC_loaded_V']:
                    try:
                        result[over] = wave.results[over]
                    except KeyError:
                        pass
                if result:
                    all_results[wave_name] = result
            
        if not verbose:
            return all_results
    
    def droop_intervals(self):

        for wave_name, wave in self.waves.items():
            logger.info(wave_name)
            wave.droop_intervals()

    def resample(self, tsample, oversample=False, verbose=True):

        resamp_waves = []
        for wave_name, wave in self.waves.items():
            resamp_waves.append(wave.resample(tsample=tsample,
                                              oversample=oversample,
                                              verbose=False
                                             )
                               )
            if verbose:
                print(wave_name)
                print(f'\tBefore resmapling: '
                      f'{len(wave.data.y)} points, after resmapling: '
                      f'{len(resamp_waves[-1].data.y)} points')
        
        return self.group(*resamp_waves)

    def filt(self, f, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        filt_waves = [wave.filt(f, tstart, tend)
                    for wave, tstart, tend in zip(self.waves.values(), ts, te)]
        return self.group(*filt_waves)

    def save(self, format_type, wave_names=None, vnom=1, sep=',',
             suffix=None, keep_original_fname=False, header=True):

        wave_names = self.waves.keys() if wave_names is None else wave_names

        logger.info(f'Saving waveforms as {format_type}:')
        for wave_name in wave_names:
            if keep_original_fname:
                fname = str(Path(self.waves[wave_name].file_name).stem)
            else:
                fname = None
            self.waves[wave_name].save(format_type=format_type,
                                       fname=fname,
                                       header=header,
                                       vnom=vnom,
                                       sep=sep, suffix=suffix)
            logger.info(f'\t{wave_name}')

    def clip(self, tstart=None, tend=None):

        ts, te = self._time_vector(tstart, tend)

        clipped_waves = [wave.clip(tstart, tend)
                        for wave, tstart, tend in zip(self.waves.values(), ts, te)]
        return self.group(*clipped_waves)

    def shift(self, tshift=None):

        shifted_waves = [wave.shift(tshift) for wave in self.waves.values()]
        return self.group(*shifted_waves)

    def extend(self, ext_time, preamble_start=None, preamble_end=None,
               data_start=None, data_end=None, postamble_start=None,
               postamble_end=None):

        extended_waves = [wave.extend(ext_time, preamble_start,
                                      preamble_end,
                                      data_start, data_end,
                                      postamble_start, postamble_end
                                     ) for wave in self.waves.values()]
        return self.group(*extended_waves)

    def abs(self):

        abs_waves = [wave.abs() for wave in self.waves.values()]

        return self.group(*abs_waves)

    def angle(self):

        angle_waves = [wave.angle() for wave in self.waves.values()]

        return self.group(*angle_waves)

    def real(self):

        real_waves = [wave.real() for wave in self.waves.values()]

        return self.group(*real_waves)
        
    def imag(self):

        imag_waves = [wave.imag() for wave in self.waves.values()]

        return self.group(*imag_waves)

    def trans_func(self, net, zs=0, zl=0, load_port=2, source_port=1):

        z0l = net.z0[0, load_port-1]
        z0s = net.z0[0, source_port-1]
        gamma_l = (zl - z0l)/(zl + z0l)
        gamma_s = (zs - z0s)/(zs + z0s)
        s11 = net.s[:, source_port-1, source_port-1]
        s12 = net.s[:, source_port-1, load_port-1]
        s21 = net.s[:, load_port-1, source_port-1]
        s22 = net.s[:, load_port-1, load_port-1]
        gamma_in = s11 + s12*s21*gamma_l/(1 - s22*gamma_l)
        tf = (s21*(1 + gamma_l)*(1 - gamma_s))/(2*(1 - s22*gamma_l)*(1 - gamma_in*gamma_s))

        prim.Wave.wave_num += 1
        return prim.Wave(DataVector(x=net.f, y=tf,
                                   x_unit='Hz',
                                   y_unit='',
                                   wave_name=f'Voltage_tf{prim.Wave.wave_num-1}',
                                   file_name=self.file_name,
                                   path=self.path,
                                   proc_hist=[f'Transfer function between load port '
                                                f'{load_port} and source port {source_port}']
                                  )
                        )

    def fft(self, n=None, win_type=None):

        fft_waves = []
        complex_waves = []
        for wave in self.waves.values():
            fft_wave, complex_wave = wave.fft(n, win_type, False)
            fft_waves.append(fft_wave)
            complex_waves.append(complex_wave)

        return self.group(*(fft_waves + complex_waves))

    def ifft(self, n=None):      

        ifft_waves = [wave.fft(n, False) for wave in self.waves.values()]

        return self.group(*ifft_waves)

    def create_BIB(self, freq, duty=50, idle=None, preamble_start=None, preamble_end=None,
                   data_start=None, data_end=None, postamble_start=None, postamble_end=None):

        BIB_waves = [wave.create_BIB(freq, duty=duty, idle=idle,
                                        preamble_start=preamble_start,
                                        preamble_end=preamble_end,
                                        data_start=data_start, data_end=data_end,
                                        postamble_start=postamble_start,
                                        postamble_end=postamble_end
                                     ) for wave in self.waves.values()]
        
        return self.group(*BIB_waves)

    def tune_icct(self, preamble_start, preamble_end, postamble_start,
                    postamble_end, freqs, dutys, vnom,
                    data_start=None, data_end=None,
                    keep_original_fname=False, include_header=True,
                    exclude=None, export_formats='gpoly',
                    idle=None):
        '''Tunes given icc(t) current waveforms to specific frequencies
        and duty cycles based on preamble and postamble definition.

        :param preamble_start: Starting time, in seconds, of the preamble window
        :type preamble_start: float
        :param preamble_end: Ending time, in seconds, of the preamble window
        :type preamble_end: float
        :param postamble_start: Starting time, in seconds, of the postamble window
        :type postamble_start: float
        :param postamble_end: Ending time, in seconds, of the preamble window
        :type postamble_end: float
        :param freqs: A list of frequencies, in Hz, to create the BIB pattern
        :type freqs: list[float]
        :param dutys: A list of duty cycles, in precentage, to create the BIB pattern.
        These numbers inidicate the high level portion of the BIB pattern.
        The length of this list has to be identical to the length of 'freqs'
        :type dutys: list[float]
        :param vnom: Nominal voltage to create the gpoly Hspice subcircuit
        :type vnom: float
        :param data_start: Starting time, in seconds, of the data window, defaults to None
        :type data_start: float, optional
        :param data_end: Ending time, in seconds, of the data window, defaults to None
        :type data_end: float, optional
        :param keep_original_fname: If True, indicating to save the tuned icc(t)
        using the original file name. If False will use the waveform name as the file name, defaults to False
        :type keep_original_fname: bool, optional
        :param include_header: If True and saving in csv format, indicating to save the tuned icc(t)
        with header names. If False and saving in csv format, header title will not be used, defaults to True
        :type include_header: bool, optional
        :param exclude: A list of waveform names to exclude from the process
        of creating a BIB pattern. Instead, the excluded waveform will be saved without changes,
        defaults to None
        :type exclude: list[str], optional
        :param export_formats: A list or string of file export formats.
        The only acceptable values are: 'gpoly', 'csv', and 'PWL', defaults to 'gpoly'
        :type export_formats: list[str] or str, optional
        :param idle: The value of the idle portion of the waveform.
        If None, uses the first point as the idle values. Otherwise a tuple of 2 floats should be provided,
        indicating the start and end time window. An average value of idle is calculated based on the
        waveform amplitude within this time window per each waveform, defaults to None
        :type idle: None or tuple[float, float], optional
        :return: BIB patterns of all combinations of frequencies and duty cycles
        :rtype: dict[:class:`primitives.Wave()`]
        '''

        if data_start is None or data_end is None:
            data_start = preamble_end
            data_end = postamble_start

        if isinstance(export_formats, str):
            export_formats = [export_formats]

        if exclude is None:
            subgroup = self
            subgroup_names = self.wave_names(verbose=False)
        else:
            subgroup_names = [wave_name for wave_name in self.waves.keys()
                                if wave_name not in exclude]
            subgroup = self.group(*[wave for wave in self.waves.values()
                                if wave.wave_name not in exclude]
                            )
            # Save the excluded waveforms without any changes
            excluded_group = self.group(*[self.waves[wave_name] for wave_name in exclude])
            for export_format in export_formats:
                excluded_group.save(format_type=export_format, vnom=vnom, suffix=export_format)

        cases = {}
        for freq, duty in zip(freqs, dutys):
            bib_waves = subgroup.create_BIB(freq=freq, duty=duty, idle=idle,
                                            preamble_start=preamble_start,
                                            preamble_end=preamble_end,
                                            data_start=data_start,
                                            data_end=data_end,
                                            postamble_start=postamble_start,
                                            postamble_end=postamble_end)
            
            # Modify the names of saved waveforms
            for (bib_name, bib_wave), prev_name in zip(bib_waves.waves.copy().items(), subgroup_names):
                bib_wave.wave_name = prev_name
                bib_waves.waves[prev_name] = bib_wave
                del bib_waves.waves[bib_name]
            cases[f'{freq*1e-6}MHz_{duty}%'] = bib_waves

            # Save waveforms in the specified formats
            for export_format in export_formats:
                if keep_original_fname:
                    suffix = None
                else:
                    suffix=f'{freq*1e-6}MHz_{duty}%_{export_format}'
                bib_waves.save(format_type=export_format,
                                keep_original_fname=keep_original_fname,
                                header=include_header,
                                vnom=vnom,
                                suffix=suffix)

        return cases
    
    def resonance_tuning(self, low2high_start, high2low_start,
                         freq, duty=50):
        
        tuned_waves = []
        for wave_name, wave in self.waves.items():
            logger.info(wave_name)
            tuned_waves.append(wave.resonance_tuning(low2high_start, high2low_start,
                                                    freq, duty))
            
        return self.group(*tuned_waves)

    def create_BIB_DDR(self, preamble_start, preamble_end, postamble_start,
                       postamble_end, freq, duty=50, num_bits=1, bits_per_packet=8):

        BIB_waves = []
        for wave_name, wave in self.waves.items():
            logger.info(wave_name)
            BIB_waves.append(wave.create_BIB_DDR(preamble_start, preamble_end,
                                                 postamble_start, postamble_end,
                                                 freq, duty=duty, num_bits=num_bits,
                                                 bits_per_packet=bits_per_packet
                                                )
                            )
        return self.group(*BIB_waves)
        
    def denoise(self, noise_att=11, reduce=1, verbose=True):

        denoised_waves = []
        for wave_name, wave in self.waves.items():
            logger.info(wave_name)
            denoised_waves.append(wave.denoise(noise_att, reduce, verbose=False))

            if verbose:
                print(f'\tBefore denoising: '
                        f'{len(wave.data.y)} points, after denoising: '
                        f'{len(denoised_waves[-1].data.y)} points')

        return self.group(*denoised_waves)

    def create_excel(self, waves):

        # Make maps and save results to an Excel file
        wb = Workbook()
        worksheets = {}
        ws_ports = wb.active
        ws_ports.title = 'Ports'

        for sheet in self.sheets:
            worksheets[sheet] = wb.create_sheet(sheet)
            worksheets[sheet]['A1'] = 'VR tol.'
            worksheets[sheet]['A1'].alignment = Alignment(horizontal='center')
            worksheets[sheet]['A1'].font = Font(bold=True)
            worksheets[sheet]['B1'] = 0
            worksheets[sheet]['B1'].alignment = Alignment(horizontal='center')
            worksheets[sheet]['D1'] = 'Min'
            worksheets[sheet]['D1'].font = Font(bold=True)
            worksheets[sheet]['D1'].alignment = Alignment(horizontal='right')
            worksheets[sheet]['G1'] = 'Max'
            worksheets[sheet]['G1'].font = Font(bold=True)
            worksheets[sheet]['G1'].alignment = Alignment(horizontal='right')

        min_row = np.inf
        min_col = np.inf
        max_row = 0
        max_col = 0
        del_col = 1
        del_row = 3
        for port_name, (row, col) in self.port_loc.items():
            min_row = min(min_row, row+del_row)
            min_col = min(min_col, col+del_col)
            max_row = max(max_row, row+del_row)
            max_col = max(max_col, col+del_col)
            ws_ports.cell(row+del_row, col+del_col, port_name)
            ws_ports.cell(row+del_row, col+del_col).alignment = Alignment(horizontal='center')

            if isinstance(waves, pd.DataFrame):
                for measr, ws in worksheets.copy().items():
                    #try:
                    if 'sw' in port_name.lower():
                        ws.cell(row+del_row, col+del_col, port_name)
                    else:
                        if 'pk2pk' in measr:
                            ws.cell(row+del_row, col+del_col, f'={waves[measr][port_name]}+2*B1')
                        elif 'min' in measr:
                            ws.cell(row+del_row, col+del_col, f'={waves[measr][port_name]}-B1')
                        else:
                            ws.cell(row+del_row, col+del_col, f'={waves[measr][port_name]}+B1')
                    ws.cell(row+del_row, col+del_col).alignment = Alignment(horizontal='center')
                    #except KeyError:
                    #    print(f'{port_name} is out of bounds')
            else:
                wave = waves[self.name_conv[port_name]]
                for result, ws in worksheets.copy().items():
                    try:
                        if result == 'pk2pk':
                            ws.cell(row+del_row, col+del_col, f'={wave.results[result][1]}+2*B1')
                        elif 'min' in result or result == 'min':
                            ws.cell(row+del_row, col+del_col, f'={wave.results[result][1]}-B1')
                        else:
                            ws.cell(row+del_row, col+del_col, f'={wave.results[result][1]}+B1')
                        ws.cell(row+del_row, col+del_col).alignment = Alignment(horizontal='center')
                    except KeyError:
                        ws.cell(row+del_row, col+del_col, port_name)
                        ws.cell(row+del_row, col+del_col).alignment = Alignment(horizontal='center')

        # Insert formulae into worksheets and apply conditional color scheme
        rule_min = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFF8696B', \
                                  mid_type='percentile', mid_value=50, mid_color='FFFFEB84', \
                                  end_type='percentile', end_value=90, end_color='FF63BE7B')
        rule_max = ColorScaleRule(start_type='percentile', start_value=10, start_color='FF63BE7B', \
                                  mid_type='percentile', mid_value=50, mid_color='FFFFEB84', \
                                  end_type='percentile', end_value=90, end_color='FFF8696B')

        tab_names = []
        for result, ws in worksheets.copy().items():
            ws['E1'] = f'=MIN({get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row})'
            ws['E1'].alignment = Alignment(horizontal='left')
            ws['H1'] = f'=MAX({get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row})'
            ws['H1'].alignment = Alignment(horizontal='left')
            if '_droop' in result or 'min' in result:
                ws.conditional_formatting.add(f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}', rule_min)
            else:
                ws.conditional_formatting.add(f'{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}', rule_max)

            tab_name = result
            for inv_char, rep_char in [('{', ''), ('}', ''), ('[', ''),
                                        (']', ''), (',', '_'), (' ', '_'),
                                        ('.', 'p')]:
               tab_name = tab_name.replace(inv_char, rep_char)
            tab_names.append(tab_name)

            new_range = openpyxl.workbook.defined_name.DefinedName(f'{tab_name}', attr_text=f"'{result}'!{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}")
            #wb.defined_names.append(new_range)
            wb.defined_names[f'{tab_name}'] = new_range
            

        for ws in worksheets.copy().values():
            for col in range(min_col, max_col+1):
                for row in range(min_row, max_row+1):
                    if ws.cell(row=row, column=col).value is None:
                        ws.cell(row, col, ' ')
        
        ws_summary = wb.create_sheet('Summary', 0)
        ws_summary[f'{get_column_letter(min_col)}{min_row}'] = "=INDIRECT(A2)"
        ws_summary['A1'] = 'Case'
        ws_summary['A1'].font = Font(bold=True)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        ws_summary['A2'] = 'Select a case'
        ws_summary.column_dimensions['A'].width = len(ws_summary['A2'].value)
        ws_summary['B1'] = 'Min'
        ws_summary['B1'].font = Font(bold=True)
        ws_summary['B1'].alignment = Alignment(horizontal='right')
        ws_summary['C1'] = f'=MIN({get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row})'
        ws_summary['C1'].alignment = Alignment(horizontal='left')
        ws_summary['D1'] = 'Max'
        ws_summary['D1'].font = Font(bold=True)
        ws_summary['D1'].alignment = Alignment(horizontal='right')
        ws_summary['E1'] = f'=MAX({get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row})'
        ws_summary['E1'].alignment = Alignment(horizontal='left')
        ws_summary['F1'] = 'Avg'
        ws_summary['F1'].font = Font(bold=True)
        ws_summary['F1'].alignment = Alignment(horizontal='right')
        ws_summary['G1'] = f'=AVERAGE({get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row})'
        ws_summary['G1'].alignment = Alignment(horizontal='left')
        ws_summary.conditional_formatting.add(f'A2:{get_column_letter(max_col)}{max_row}', rule_max)
        for col in range(min_col, max_col+1):
            for row in range(min_row, max_row+1):
                ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal='center') 

        for idx, tab_name in enumerate(tab_names):
            ws_ports[f'A{max_row+2+idx}'] = tab_name
            ws_ports.row_dimensions[max_row+2+idx].hidden = True

        data_val = DataValidation(type="list", formula1=f'=Ports!$A${max_row+2}:$A${max_row+2+len(tab_names)}')
        ws_summary.add_data_validation(data_val)
        data_val.add(ws_summary["A2"])
        
        return wb

    def _heatmap_setup(self, port_map_file, port_conv_file, match=True):

        port_map = pd.read_csv(port_map_file).to_numpy()

        # Make port location map
        for x in range(port_map.shape[0]):
            for y in range(port_map.shape[1]):
                if isinstance(port_map[x, y], str):
                    self.port_loc[port_map[x, y]] = (x, y)

        # Make port conversion map
        if port_conv_file is not None:
            name_map = pd.read_excel(port_conv_file, header=None)
            self.name_conv = {port_name:wave_name for wave_name, port_name in zip(name_map[0], name_map[1])}
        elif match: # Find matches between waveform names and port names
            wave_names = self.wave_names(verbose=False)
            for port_name in self.port_loc.keys():
                for cutoff in np.arange(1, 0, -0.1):
                    matches = get_close_matches(port_name, wave_names, n=1, cutoff=cutoff)
                    if matches:
                        self.name_conv[port_name] = matches[0]
                        logger.info(f'{port_name}\tis matched to\t{matches[0]}\twith {cutoff:.3f} confidence level')
                        del wave_names[wave_names.index(matches[0])]
                        break
                if not matches:
                    logger.warning(f'Cannot find a match to a {port_name}')
            if wave_names:
                logger.warning('\nCould not find matches to the following waveforms:')
                for wave_name in wave_names:
                    logger.info(wave_name)
    
    def heatmap_filter(self, filters, port_map_file, port_conv_file=None, tstart=None, tend=None):

        self._heatmap_setup(port_map_file, port_conv_file)

        filt_waves = {}
        if not isinstance(filters, list):
            filters = [filters]

        self.sheets = []
        for filt in filters:
            freq = filt._freq_descr()
            self.sheets = self.sheets + [f'min {freq}', f'max {freq}', f'pk2pk {freq}']

            for wave_name, wave in self.waves.items():
                filt_wave = wave.clip(tstart, tend).filt(filt)
                _ = filt_wave.pk2pk()
                for result in ['min', 'max', 'pk2pk']:
                    filt_wave.results[f'{result} {freq}'] = filt_wave.results.pop(result)
                if wave_name in filt_waves:
                    filt_waves[wave_name].results = {**filt_waves[wave_name].results, **filt_wave.results}
                else:
                    filt_waves[wave_name] = filt_wave

        wb = self.create_excel(filt_waves)
        wb.save(os.path.join(self.path, 'reports', 'heatmap_filter.xlsx'))

    def heatmap_trans(self, port_map_file, port_conv_file=None, tstart=None, tend=None):

        self._heatmap_setup(port_map_file, port_conv_file)

        self.sheets = ['1st_droop', '2nd_droop', '3rd_droop',
                       '1st_over', '2nd_over', '3rd_over',
                       '1st_min', '2nd_min', '3rd_min',
                       '1st_max', '2nd_max', '3rd_max',
                       'min', 'max', 'pk2pk', 'DC_loaded']
        for wave in self.waves.values():
            wave.find_droops(tstart, tend)
            wave.find_overshoots(tstart, tend)
            _ = wave.pk2pk(tstart, tend)

        wb = self.create_excel(self.waves)
        wb.save(os.path.join(self.path, 'reports', 'heatmap_trans.xlsx'))

    def heatmap_vec(self, port_map_file, vec_file, port_conv_file=None, heatmap_fname=None):

        self._heatmap_setup(port_map_file, port_conv_file, match=False)

        if isinstance(vec_file, str):
            port_map = pd.read_csv(vec_file, index_col=0)
        else:
            port_map = vec_file

        self.sheets = port_map.columns
        wb = self.create_excel(port_map)
        heatmap_fname = 'heatmap_vec.xlsx' if heatmap_fname is None else heatmap_fname
        try:
            wb.save(os.path.join(self.path, 'reports', heatmap_fname))
        except FileNotFoundError:
            wb.save(os.path.join(self.path, heatmap_fname))

    def heatmap_mat(self, port_map_file, mat_file, port_conv_file=None):

        self._heatmap_setup(port_map_file, port_conv_file, match=False)

        port_map = pd.read_csv(mat_file, header=None).dropna(how='all').reset_index(drop=True)
        x_len = port_map.shape[1] + 1
        y_len = port_map.shape[0]

        sub_mats = []
        for idx in range(y_len//x_len):
            sub_mat = port_map[idx*x_len:idx*x_len+x_len].reset_index(drop=True)
            headers = sub_mat.iloc[0]
            sheet_title = 'NoName'
            for header in headers:
                if isinstance(header, str):
                    sheet_title = header
                    break
            sub_mat = sub_mat.drop(0)
            sub_mats.append(pd.DataFrame(np.diag(sub_mat), columns=[sheet_title],
                            dtype=float))
        port_map = pd.concat(sub_mats, axis=1)

        self.sheets = port_map.columns
        wb = self.create_excel(port_map)
        wb.save(os.path.join(self.path, 'reports', 'heatmap_mat.xlsx'))

    def heatmap_ind_res(self, port_map_file, freq, short_name='sw',
                        net=None, port_conv_file=None):

        self._heatmap_setup(port_map_file, port_conv_file, match=False)

        if net is None:
            # Find and short all SW ports
            group = self._auto_term(short_name)
        else:
            group = self
        idx_freq, net = group.get_net_param(freq)
        
        ind_res = pd.DataFrame({f'L nH {int(freq*1e-6)} MHz': np.diag(net.z_im[idx_freq])/(2*np.pi*freq)*1e9,
                                 f'R mOhm {int(freq*1e-6)} MHz': np.diag(net.z_re[idx_freq]*1e3),
                                 'Rdc mOhm': np.diag(net.z_re[0])*1e3},
                                 index=[wave_name for wave_name in self.wave_names(verbose=False)
                                        if 'sw' not in wave_name.lower()]
                                )

        self.sheets = ind_res.columns
        wb = self.create_excel(ind_res)
        wb.save(os.path.join(self.path, 'reports', 'heatmap_ind_res.xlsx'))

    def heatmap_ports(self, freq, from_ports, exclude_ports=None):

        freq_idx = self.get_net_param(freq)[0]
        ignore_port_names = self.port_selection(*exclude_ports) \
                            if exclude_ports is not None else []
        port_names = self.port_selection(*from_ports)
        to_port_names = [port_name for port_name in self.waves.keys()
                            if port_name not in port_names
                    ]

        ind_matrix = {}
        res_matrix = {}
        logger.info('Calculating loop inductance and resistance for these locations:')
        for out_name in to_port_names:
            if out_name in ignore_port_names:
                continue
            logger.info(f'\t{out_name}')
            terms = [('open', [other for other in to_port_names if out_name != other]),
                        ('short', [out_name])]
            new_net = self._terminate(*terms)[1]
            ind_matrix[out_name] = np.diag(new_net.z_im[freq_idx]/(2*np.pi*freq))*1e9
            res_matrix[out_name] = np.diag(new_net.z_re[freq_idx])*1e3

        df_ind = pd.DataFrame(ind_matrix, index=port_names)
        df_ind.index.name = 'from_port'
        df_ind.columns.name = 'to_port'
        df_ind = pd.DataFrame(df_ind.stack(), columns=['Inductance']).reset_index()
        df_res = pd.DataFrame(res_matrix, index=port_names)
        df_res.index.name = 'from_port'
        df_res.columns.name = 'to_port'
        df_res = pd.DataFrame(df_res.stack(), columns=['Resistance']).reset_index()
        
        all_plots = []
        output_file(os.path.join(self.path, 'plots', f'cap_heatmap.html'))
        port_names.reverse()
        all_plots.append(self.plt.plot_heatmap(df_ind, [out_name for out_name in to_port_names
                                                        if out_name not in ignore_port_names
                                                        ],
                                                port_names, 'nH',
                                                f'Loop Inductance [nH] '
                                                f'at {freq*1e-6:.2f} MHz'
                                                )
                        )
        all_plots.append(self.plt.plot_heatmap(df_res, [out_name for out_name in to_port_names
                                                        if out_name not in ignore_port_names
                                                        ],
                                                port_names, 'mOhm',
                                                f'Loop Resistance [mOhm] '
                                                f'at {freq*1e-6:.2f} MHz'
                                                )
                        )

        show(column(all_plots))

        return df_ind, df_res

    def create_BIB_IO(self, cfg, raw_only=False, save_as='gpoly'):

        reload(cfg)

        self.load_waves(cfg.raw_waveforms_folder)
        dashboard = pd.read_excel(cfg.dashboard_file)
        DashInfo = namedtuple('DashInfo', 'ip_name count fname gb shift idle tstart tend')

        tmin = np.inf
        info = {}
        for idx in range(len(dashboard)):
            info[dashboard.iloc[idx, 0].replace(' ', '_')] = DashInfo(ip_name = dashboard.iloc[idx, 0].replace(' ', '_'),
                                                                      count = dashboard.iloc[idx, 1],
                                                                      fname = dashboard.iloc[idx, 2],
                                                                      gb = dashboard.iloc[idx, 3],
                                                                      shift = dashboard.iloc[idx, 4]*1e-12,
                                                                      idle = dashboard.iloc[idx, 5]*1e-3,
                                                                      tstart = dashboard.iloc[idx, 6]*1e-9,
                                                                      tend = dashboard.iloc[idx, 7]*1e-9)
            tmin = min(tmin, info[dashboard.iloc[idx, 0].replace(' ', '_')].shift)

        stats = pd.DataFrame(columns=['Name', 'Count', 'Guardband [%]', 'Offset [ps]', 'Idle Icc(t) [mA]',
                                      'Iavg per unit [mA]', 'Irms per unit [mA]', 'Iavg per IP [mA]',
                                      'Irms per IP [mA]', 'Iavg per IP+GB [mA]', 'Irms per IP+GB [mA]'])
        fmap = {wave.file_name:wave for wave in self.waves.values()}
        new_waves = {}
        for ip_name, ip_info in info.items():
            new_waves[ip_name] = deepcopy(fmap[ip_info.fname])
            new_waves[ip_name].wave_name = ip_name
            new_waves[ip_name].y_unit = 'A'
            Iavg = np.mean(new_waves[ip_name].data.y)*1e3
            Irms = np.sqrt(np.mean(new_waves[ip_name].data.y**2))*1e3
            stats.loc[len(stats.index)] = [ip_name, ip_info.count,
                                            ip_info.gb*100, ip_info.shift*1e12,
                                            ip_info.idle*1000, Iavg, Irms,
                                            Iavg*ip_info.count, Irms*ip_info.count,
                                            Iavg*ip_info.count*(1+ip_info.gb),
                                            Irms*ip_info.count*(1+ip_info.gb)]
        self.waves = new_waves

        #print('\nSummary Table saved in icct_stats.xlsx')
        #print('-'*len('Summary Table saved in icct_stats.xlsx')+'\n')
        #pd.set_option('display.max_columns', 20)
        #pd.set_option('display.width', 1000)
        #print(stats)
        stats.to_excel(os.path.join(self.path, 'reports', 'icct_stats.xlsx'))

        # Plot all stacked raw wavewforms
        logger.info('\nPlotting raw icc(t)... ', end='')
        self.plot_stack(y_scale='m', out_file=f'raw_icct{cfg.suffix}')
        logger.info('Done')

        if raw_only:
            return

        # Calculate and plot individual icc(t) BIB
        single_BIBs = []
        for freq, duty in zip(cfg.BIB_frequency, cfg.burst_duty_cycle):
            suffix = f'{freq*1e-6}MHz_{duty}'.replace('.', 'p')
            for wave_name, wave in self.waves.items():
                try:
                    BIB_wave = wave*info[wave_name].count*(1 + info[wave_name].gb)*cfg.mul
                    BIB_wave = BIB_wave.create_BIB(
                                        freq=freq, duty=duty,
                                        idle=info[wave_name].idle*info[wave_name].count*(1 + info[wave_name].gb)*cfg.mul,
                                        data_start=info[wave_name].tstart,
                                        data_end=info[wave_name].tend).shift(info[wave_name].shift-tmin
                                                    )
                    BIB_wave.wave_name = f'{wave_name}_{suffix}'
                    single_BIBs.append(BIB_wave)
                except KeyError:
                    pass
            logger.info(f"Plotting BIB icc(t) {suffix.replace('p', '.').replace('_', ' ')}%... ", end='')
            self.plot_stack([bib for bib in single_BIBs if suffix in bib.wave_name],
                            y_scale='m', out_file=f'BIB_icct{cfg.suffix}_{suffix}')
            logger.info('Done')

        # Save all BIBs
        logger.info('Saving the following BIB waveforms:')
        for bib in single_BIBs:
            logger.info(f'\t{bib.wave_name}')
            bib.save(save_as, fname=f'{bib.wave_name}{cfg.suffix}', vnom=cfg.vnom)
        logger.info('Done')

        '''
        # Calculate and plot total icc(t) BIB
        total_BIBs = []
        for freq, duty in zip(cfg.BIB_frequency, cfg.burst_duty_cycle):
            suffix = f'{freq*1e-6}MHz_{duty}'.replace('.', 'p')
            group_BIB = []
            for wave in single_BIBs:
                if suffix in wave.wave_name:
                    group_BIB.append(wave)
            total_BIB = sum(group_BIB)
            total_BIB.wave_name = f'total{cfg.suffix}_{suffix}'
            total_BIB.data.y[0] = total_BIB.data.y[-1]
            total_BIBs.append(total_BIB)
        print('Plotting total BIB icc(t)... ', end='')
        self.plot_stack(total_BIBs, y_scale='m', clr='firebrick', out_file=f'total_BIB_icct{cfg.suffix}')
        print('Done')

        # Save total icc(t) BIB
        for freq, duty in zip(cfg.BIB_frequency, cfg.burst_duty_cycle):
            suffix = f'{freq*1e-6}MHz_{duty}'.replace('.', 'p')
            for wave in total_BIBs:
                if suffix in wave.wave_name:
                    print(f"Saving total BIB icc(t) {suffix.replace('p', '.').replace('_', ' ')}%... ", end='')
                    wave.save(save_as, fname=f'total_BIB{cfg.suffix}_{suffix}', vnom=cfg.vnom)
                    print('Done')
        '''

    def calc_AVP_VID(self, v_spec, vr_tol, iload):

        vmin_spec = v_spec[0] + vr_tol[0]
        vmax_spec = v_spec[1] - vr_tol[1]
        imin, imax = iload

        # Find max, min, and min of DC_loaded
        vmin = np.inf
        vmax = -np.inf
        for wave in self.waves.values():
            wave.find_overshoots()
            wave.find_droops()
            if wave.results['min'][1] < vmin:
                vmin = wave.results['min'][1]
                DC_unloaded = wave.results['DC_unloaded'][1]
            if wave.results['max'][1] > vmax:
                vmax = wave.results['max'][1]
                DC_loaded = wave.results['DC_loaded'][1]
        droop = DC_unloaded - vmin
        overshoot = vmax - DC_loaded
        if vmin >= vmin_spec and vmax <= vmax_spec:
            logger.info(f'Spec is already satisfied with {(vmin - vmin_spec)*1000:.3f} mV '
                    f'min margin and {(vmax_spec - vmax)*1000:.3f} mV max margin')
            return

        # Solving for two vmin constraints
        AVP_min = droop/(imax - imin)
        VID_min = vmin_spec - DC_unloaded + (imax*droop)/(imax - imin)

        # Solving for two vmax constraints
        AVP_max = overshoot/(imax - imin)
        VID_max = vmax_spec - DC_loaded + (imin*overshoot)/(imax - imin)

        logger.info(f'AVP range is {AVP_min*1000:.3f} mΩ to {AVP_max*1000:.3f} mΩ')
        logger.info(f'VID shift range is {VID_min*1000:.3f} mV to {VID_max*1000:.3f} mV')
        logger.info(f'Mid point is (VID, AVP) = '
                    f'({(VID_min + VID_max)/2*1000:.3f} mV, ' 
                    f'{(AVP_min + AVP_max)/2*1000:.3f} mΩ)')

    def pow(self, n):

        pow_waves = [wave.pow(n) for wave in self.waves.values()]
        return self.group(*pow_waves)

    def mean_sqaure_error(self, other_waves, tstart=None, tend=None, verbose=True):

        mse = []
        for self_wave, other_wave in zip(self.waves.values(),
                                        other_waves.waves.values()):
            mse.append(self_wave.mean_square_error(other_wave, tstart, tend))

        max_err = np.max(mse)
        min_err = np.min(mse)
        avg_err = np.mean(mse)
        spread_err = max_err - min_err
        if verbose:
            print(f'Maximum error: {max_err}')
            print(f'Minimum error: {min_err}')
            print(f'Average error: {avg_err}')
            print(f'Error spread: {spread_err}')
        else:
            return {'max_err': max_err, 'min_err': min_err,
                    'avg_err': avg_err, 'spread_err': spread_err}

    def detect_settle(self):
        '''Finds a settling time window of the group waveforms.
        Settling is assumed to accure at the end of the waveform.

        :return: Settle time window of each waveform in the group
        :rtype: dict
        '''

        return {wave_name: wave.detect_settle()
                for wave_name, wave in self.waves.items()}

    def detect_change_points(self, plot_fig=True, slew_perc=None):
        '''Detects change points in the x-axis domain.
        This algorithm assumes one cycle of the periodic waveform
        as well as up to three distinguished regions.

        :param plot_fig: If True, indiciating to generate a plot with
        a marked time window, defaults to True
        :type plot_fig: bool, optional
        :return: Change points and the corresponding statistics
        :rtype: dict of DataFrame or dict of tuples depending on save_stats
        '''

        region_statistics = {}
        legend = {'Waveform Name': [], 'File Name': []}
        for wave_name, wave in self.waves.items():
            region_statistics[wave_name] = wave.detect_change_points(plot_fig, slew_perc)
            legend['Waveform Name'].append(wave_name)
            legend['File Name'].append(wave.file_name)

        # Save stat tables as seperate tabs
        if plot_fig:
            region_statistics['Legend'] = pd.DataFrame(legend)
            with pd.ExcelWriter(os.path.join(self.path, 'reports', 'waveform_stats.xlsx')) as writer:
                for sheet_name, stat in region_statistics.items():
                    stat.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return region_statistics

    def detect_slew_rate(self, slew_perc=90):
        '''Automatically detects slew rate (A/sec) of all waveforms. 

        :param slew_perc: Percent of the amplitude to use
        when calculating the slew rate. For example, 90 means 90%-10%,
        80 means 80%-20%. One setting will be applied to all
        waveforms in the group, defaults to 90 percent.
        :type slew_perc: int, optional
        :return: Dictionary with the waveform names as keys and
        slew rate, start time, and end time as the values
        :rtype: dict[tuple]
        '''

        return {wave_name:wave.detect_slew_rate(slew_perc=slew_perc)
                for wave_name, wave in self.waves.items()}

    def slew_rate(self, tstart=None, tend=None):
        '''Calculates slew rates of all waveforms and returns them in a dict.

        :param tstart: Start time of the truncation window in seconds, defaults to None
        :type tstart: float, None, optional
        :param tend: End time of the truncation window in seconds, defaults to None
        :type tend: float, None, optional
        :return: Slew rate of the waveform
        :rtype: dict
        '''

        ts, te = self._time_vector(tstart, tend)

        return {wave_name:wave.slew_rate(tstart, tend)
                for (wave_name, wave), tstart, tend in zip(self.waves.items(), ts, te)}

    def x_at_y(self, y):
        '''Finds x value at a given y value and returns them in a dict.

        :param y: Vertical axis value
        :type y: float or list
        :return: The corresponding horizontal x-axis value
        :rtype: float
        '''

        ys, _ = self._time_vector(y, None)

        return {wave_name:wave.x_at_y(y)
                for (wave_name, wave), y in zip(self.waves.items(), ys)}

    def y_at_x(self, x):
        '''Finds y value at a given x value and returns them in a dict.

        :param x: Horizontal axis value
        :type x: float or list
        :return: The corresponding vertical y-axis value
        :rtype: float
        '''
        
        xs, _ = self._time_vector(x, None)

        return {wave_name:wave.y_at_x(x)
                for (wave_name, wave), x in zip(self.waves.items(), xs)}
